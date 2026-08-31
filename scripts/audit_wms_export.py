"""Validate a generated WMS Excel against the source 1C v7.7 order."""

from __future__ import annotations

import json
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook

from src.adapters.wms_excel_adapter import (
    SUMMARY_SHEET_NAME,
    WMS_COLUMNS,
    WMS_SHEET_NAME,
    WMSExcelAdapter,
    auto_filter_data_last_row,
)
from src.parsers.v7_parser import parse_v7_order
from src.utils.logger import console

OUTPUT_DIR = PROJECT_ROOT / "output"
DEFAULT_ORDER = PROJECT_ROOT / "data" / "orders" / "order_transfering_01_09.xls"
EXPECTED_ROWS = 384
EXPECTED_QTY_SUM = 871
FORBIDDEN_BARCODE_TOKENS = {"none", "nan", "null"}
PREVIEW_LIMIT = 20


def _latest_wms_xlsx(output_dir: Path) -> Path:
    files = sorted(output_dir.glob("WMS_Импорт_*.xlsx"), key=lambda path: path.stat().st_mtime)
    if not files:
        raise FileNotFoundError(f"No WMS_Импорт_*.xlsx in {output_dir}")
    return files[-1]


def _fail(message: str) -> None:
    console.print(f"[red]FAIL[/red] {message}")
    raise SystemExit(1)


def _excel_int(value: object) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise TypeError(f"{value!r} is not int")
    return int(value)


def main() -> None:
    order_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_ORDER
    order_path = order_arg if order_arg.is_absolute() else PROJECT_ROOT / order_arg
    if not order_path.exists():
        _fail(f"Source order not found: {order_path}")

    parsed = parse_v7_order(order_path)
    source_blocks = sorted(parsed.blocks, key=lambda block: block.order_line_number)
    source_rows = len(source_blocks)
    source_qty = sum(block.quantity for block in source_blocks)

    xlsx_path = _latest_wms_xlsx(OUTPUT_DIR)
    workbook = load_workbook(xlsx_path, data_only=False)
    if workbook.sheetnames[:2] != [WMS_SHEET_NAME, SUMMARY_SHEET_NAME]:
        _fail(
            f"Unexpected sheets {workbook.sheetnames!r}, "
            f"expected [{WMS_SHEET_NAME!r}, {SUMMARY_SHEET_NAME!r}]"
        )
    worksheet = workbook[WMS_SHEET_NAME]
    try:
        data_last_row = auto_filter_data_last_row(worksheet)
    except ValueError as exc:
        _fail(str(exc))
    excel_rows = max(0, data_last_row - 1)
    expected_filter = f"A1:E{WMSExcelAdapter.wms_data_last_row(excel_rows)}"
    if worksheet.auto_filter.ref != expected_filter:
        _fail(f"Autofilter {worksheet.auto_filter.ref!r} != {expected_filter!r}")
    totals_row = WMSExcelAdapter.wms_totals_row(excel_rows)
    if int(worksheet.max_row or 0) != totals_row:
        _fail(f"Expected ИТОГО on row {totals_row}, max_row={worksheet.max_row}")
    totals_name = worksheet.cell(row=totals_row, column=2).value
    if not isinstance(totals_name, str) or not totals_name.startswith("ИТОГО"):
        _fail(f"Totals label {totals_name!r} is not ИТОГО")
    if f"Позиций: {excel_rows}" not in str(totals_name):
        _fail(f"Totals label missing position count: {totals_name!r}")
    totals_formula = worksheet.cell(row=totals_row, column=4).value
    expected_formula = f"=SUM(D2:D{data_last_row})"
    if totals_formula != expected_formula:
        _fail(f"Totals formula {totals_formula!r} != {expected_formula!r}")
    headers = [worksheet.cell(row=1, column=col).value for col in range(1, 6)]
    if headers != WMS_COLUMNS:
        _fail(f"Unexpected headers {headers!r}, expected {WMS_COLUMNS}")
    if headers[0] != "№":
        _fail(f"First header must be '№', got {headers[0]!r}")

    qty_sum = 0
    line_numbers: list[int] = []
    for row_index in range(2, data_last_row + 1):
        number_cell = worksheet.cell(row=row_index, column=1)
        barcode_cell = worksheet.cell(row=row_index, column=3)
        qty_cell = worksheet.cell(row=row_index, column=4)
        try:
            line_no = _excel_int(number_cell.value)
        except TypeError:
            _fail(f"Row {row_index}: № {number_cell.value!r} is not int")
        line_numbers.append(line_no)

        barcode_value = barcode_cell.value
        if barcode_value is not None:
            barcode_text = str(barcode_value).strip()
            if barcode_text.lower() in FORBIDDEN_BARCODE_TOKENS:
                _fail(f"Row {row_index}: barcode is the string {barcode_text!r}")
            if barcode_text and (not barcode_text.isdigit() or len(barcode_text) != 13):
                _fail(f"Row {row_index}: barcode {barcode_text!r} is not 13 digits or empty")
        qty_value = qty_cell.value
        if not isinstance(qty_value, int) or isinstance(qty_value, bool):
            _fail(f"Row {row_index}: quantity {qty_value!r} is not int")
        qty_sum += int(qty_value)

    expected_numbers = list(range(1, excel_rows + 1))
    if line_numbers != expected_numbers:
        _fail(f"Column № is {line_numbers[:8]!r}…, expected {expected_numbers[:8]!r}… through N")

    if source_rows != excel_rows:
        _fail(f"Row count mismatch: source {source_rows} != excel {excel_rows}")
    if source_qty != qty_sum:
        _fail(f"Quantity sum mismatch: source {source_qty} != excel {qty_sum}")
    if source_rows != EXPECTED_ROWS:
        _fail(f"Source rows {source_rows} != expected {EXPECTED_ROWS}")
    if source_qty != EXPECTED_QTY_SUM:
        _fail(f"Source qty sum {source_qty} != expected {EXPECTED_QTY_SUM}")

    for index, block in enumerate(source_blocks):
        excel_row = index + 2
        excel_no = worksheet.cell(row=excel_row, column=1).value
        excel_qty = worksheet.cell(row=excel_row, column=4).value
        if int(excel_no) != block.order_line_number:
            _fail(
                f"Order identity at Excel row {excel_row}: №={excel_no} "
                f"!= parse block {block.order_line_number}"
            )
        if int(excel_qty) != int(block.quantity):
            _fail(
                f"Order identity at Excel row {excel_row}: qty={excel_qty} "
                f"!= parse block {block.order_line_number} qty={block.quantity}"
            )

    preview_path = OUTPUT_DIR / "last_wms_preview.json"
    if not preview_path.exists():
        _fail(f"Missing Streamlit/export preview sidecar: {preview_path}")
    preview_payload = json.loads(preview_path.read_text(encoding="utf-8"))
    preview_rows = preview_payload.get("preview") or []
    compare_n = min(PREVIEW_LIMIT, excel_rows, len(preview_rows))
    if compare_n < min(PREVIEW_LIMIT, excel_rows):
        _fail(f"Preview has {len(preview_rows)} rows, need {min(PREVIEW_LIMIT, excel_rows)}")
    for index in range(compare_n):
        excel_row = {
            "№": int(worksheet.cell(row=index + 2, column=1).value),
            "Наименование": worksheet.cell(row=index + 2, column=2).value,
            "Штрихкод": worksheet.cell(row=index + 2, column=3).value or "",
            "Количество": int(worksheet.cell(row=index + 2, column=4).value),
            "Заказчик": worksheet.cell(row=index + 2, column=5).value,
        }
        expected = preview_rows[index]
        expected_norm = {
            "№": int(expected["№"]),
            "Наименование": expected["Наименование"],
            "Штрихкод": expected.get("Штрихкод") or "",
            "Количество": int(expected["Количество"]),
            "Заказчик": expected["Заказчик"],
        }
        if excel_row != expected_norm:
            _fail(f"Preview mismatch at Excel row {index + 2}: {excel_row} != {expected_norm}")

    summary = workbook[SUMMARY_SHEET_NAME]
    summary_blob = " ".join(
        str(summary.cell(row=row, column=col).value or "")
        for row in range(1, min(int(summary.max_row or 1), 80) + 1)
        for col in range(1, 5)
    )
    if "Карантин" not in summary_blob:
        _fail("Summary sheet is missing the quarantine block")
    if "без ШК" not in summary_blob.lower() and "Без ШК" not in summary_blob:
        _fail("Summary sheet is missing the no-barcode block")

    workbook.close()
    console.print(f"[green]PASS[/green] {xlsx_path.name}")
    console.print(
        f"Rows {excel_rows}=={source_rows}=={EXPECTED_ROWS}; "
        f"qty {qty_sum}=={source_qty}=={EXPECTED_QTY_SUM}; "
        f"№=[1..{excel_rows}]; ИТОГО {expected_formula}; "
        f"sheets={WMS_SHEET_NAME}+{SUMMARY_SHEET_NAME}; "
        f"first {compare_n} rows match Streamlit/WMS mapping."
    )


if __name__ == "__main__":
    main()
