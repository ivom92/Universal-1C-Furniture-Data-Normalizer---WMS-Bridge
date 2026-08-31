"""Synthetic 1C v7.7 chaos generator: HTML-as-XLS, shifted headers, messy dimensions."""

from __future__ import annotations

import argparse
import html
import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.models import CatalogEntity
from src.parsers.v8_loader import load_catalog_v8

CATALOG_PATH = PROJECT_ROOT / "data" / "catalog_v8.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "synthetic"
MAIN_FILL = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
ALIAS_FILL = PatternFill(start_color="FFFFC0", end_color="FFFFC0", fill_type="solid")
SAMPLE_SIZE = 30
RNG_SEED = 80


def sample_catalog_items(
    catalog: list[CatalogEntity],
    count: int = SAMPLE_SIZE,
    rng: random.Random | None = None,
) -> list[CatalogEntity]:
    rng = rng or random.Random(RNG_SEED)
    eligible = [entity for entity in catalog if (entity.nomenclature or "").strip()]
    if len(eligible) < count:
        raise ValueError("Catalog is too small to sample synthetic orders")
    return rng.sample(eligible, count)


def distort_name(name: str, rng: random.Random) -> str:
    distorted = name.replace("х", rng.choice(["*", "x", "х"]))
    distorted = distorted.replace("дуб сонома", "д.сон.")
    distorted = distorted.replace("Дуб сонома", "д.сон.")
    distorted = distorted.replace("белое дерево", "б.дер.")
    distorted = distorted.replace("ясень шимо", "яс.шимо")
    return distorted


def write_shifted_html_xls(path: Path, items: list[CatalogEntity], *, shift_rows: int = 7) -> Path:
    """HTML table saved with a .xls suffix; header shifted down; NBSP in cells."""
    nbsp = "\xa0"
    rows: list[list[str]] = []
    for _ in range(shift_rows):
        rows.append(["", "", "", "", ""])
    rows.append(["Отборочный лист", "", "", "", ""])
    rows.append([f"Покупатель:{nbsp}", f"ИП{nbsp}Хаос{nbsp}Тестов", "", "", ""])
    rows.append(["№", "Наименование", "Тип", "Кол-во", "Ед."])
    for index, entity in enumerate(items, start=1):
        name = distort_name(entity.nomenclature, random.Random(index)).replace(" ", nbsp)
        rows.append([str(index), name, "Пачка", "1", "шт"])
        rows.append(["", f"IMP{nbsp}{entity.nomenclature[:40]}", "", "", ""])
        rows.append(["", "Продажи оптовые УРП_chaos Заказ: ЦНТ-800", "", "", ""])

    body = []
    for row in rows:
        cells = "".join(f"<td>{html.escape(cell)}</td>" for cell in row)
        body.append(f"<tr>{cells}</tr>")
    markup = (
        "<html><head><meta charset='utf-8'></head><body><table>"
        + "".join(body)
        + "</table></body></html>"
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(markup, encoding="utf-8")
    return path


def write_html_xls_with_skipped_cells(path: Path, items: list[CatalogEntity]) -> Path:
    """HTML-as-XLS with empty cells, ragged rows, and still sequential № 1..N."""
    nbsp = "\xa0"
    rows: list[list[str]] = [
        ["", "", "", ""],
        ["Отборочный лист", "", "", ""],
        ["Покупатель:", f"ИП{nbsp}Хаос{nbsp}Дырки", "", ""],
        ["", "", "", ""],
        ["№", "Наименование", "Тип", "Кол-во"],
    ]
    for index, entity in enumerate(items, start=1):
        name = distort_name(entity.nomenclature, random.Random(index + 17))
        rows.append([str(index), name, "Пачка", "1", ""])
        rows.append(["", "", "", ""])
        rows.append(["", f"IMP {entity.nomenclature[:36]}", "", "", ""])
        rows.append(["", "", "Продажи оптовые УРП_chaos Заказ: ЦНТ-801", ""])

    body = []
    for row in rows:
        cells = "".join(f"<td>{html.escape(cell)}</td>" for cell in row)
        body.append(f"<tr>{cells}</tr>")
    markup = (
        "<html><head><meta charset='utf-8'></head><body><table>"
        + "".join(body)
        + "</table></body></html>"
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(markup, encoding="utf-8")
    return path


def write_messy_dimensions_xlsx(path: Path, items: list[CatalogEntity]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отборочная"
    sheet.cell(1, 1, "Отборочный лист")
    sheet.cell(2, 1, "Покупатель:")
    sheet.cell(2, 2, "ИП Хаос Тестов")
    sheet.cell(5, 1, "№")
    sheet.cell(5, 2, "Адрес")
    sheet.cell(5, 3, "Наименование")
    sheet.cell(5, 4, "Кол-во")

    row = 6
    for index, entity in enumerate(items, start=1):
        name = entity.nomenclature.replace("х", "*").replace("×", "*")
        if "д.сон" not in name.lower() and "сонома" in name.lower():
            name = name.replace("сонома", "сон.")
            name = name.replace("Дуб", "д.")
        name = f"Р10.С2.Я1 {name}"
        sheet.cell(row, 1, index)
        loc = sheet.cell(row, 2, "Р10.С2.Я1")
        loc.fill = MAIN_FILL
        desc = sheet.cell(row, 3, name)
        desc.fill = MAIN_FILL
        sheet.cell(row, 4, 1)
        alias = sheet.cell(row + 1, 3, entity.nomenclature)
        alias.fill = ALIAS_FILL
        sheet.cell(row + 2, 3, "Продажи оптовые УРП_chaos Заказ: ЦНТ-800")
        row += 3

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def write_no_aliases_xls(path: Path, items: list[CatalogEntity]) -> Path:
    import xlwt
    from xlwt.Style import add_palette_colour

    add_palette_colour("v7_green", 0x21)
    workbook = xlwt.Workbook()
    workbook.set_colour_RGB(0x21, 0xE0, 0xFF, 0xE0)
    sheet = workbook.add_sheet("Отборочная")
    green = xlwt.easyxf("pattern: pattern solid, fore_colour v7_green")

    sheet.write(0, 0, "Отборочный лист")
    sheet.write(1, 0, "Покупатель:")
    sheet.write(1, 1, "ИП Хаос Тестов")
    sheet.write(4, 0, "№")
    sheet.write(4, 1, "Наименование")
    sheet.write(4, 2, "Кол-во")

    row = 5
    for index, entity in enumerate(items, start=1):
        sheet.write(row, 0, index)
        sheet.write(row, 1, entity.nomenclature, green)
        sheet.write(row, 2, 1)
        sheet.write(row + 1, 1, "Продажи оптовые УРП_chaos Заказ: ЦНТ-800")
        row += 2

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
    return path


def write_chaos_variant(
    path: Path,
    items: list[CatalogEntity],
    *,
    shift_rows: int,
    as_html: bool,
    skip_aliases: bool,
    star_dimensions: bool,
) -> Path:
    """One of 20+ layout mutations used by the chaos suite."""
    subset = items[: max(3, min(len(items), 8))]
    if as_html:
        return write_shifted_html_xls(path.with_suffix(".xls"), subset, shift_rows=shift_rows)
    if skip_aliases:
        return write_no_aliases_xls(path.with_suffix(".xls"), subset)
    if star_dimensions:
        return write_messy_dimensions_xlsx(path.with_suffix(".xlsx"), subset)
    return write_messy_dimensions_xlsx(path.with_suffix(".xlsx"), subset)


def generate_named_fixtures(output_dir: Path, catalog: list[CatalogEntity] | None = None) -> list[Path]:
    catalog = catalog or load_catalog_v8(CATALOG_PATH)
    items = sample_catalog_items(catalog)
    output_dir.mkdir(parents=True, exist_ok=True)
    return [
        write_shifted_html_xls(output_dir / "test_order_shifted_html.xls", items, shift_rows=7),
        write_messy_dimensions_xlsx(output_dir / "test_order_messy_dimensions.xlsx", items),
        write_no_aliases_xls(output_dir / "test_order_no_aliases.xls", items),
        write_html_xls_with_skipped_cells(
            output_dir / "test_order_skipped_cells.xls",
            items[:8],
        ),
    ]


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate chaotic synthetic 1C v7.7 order files.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    paths = generate_named_fixtures(args.output_dir)
    for path in paths:
        print(path)


if __name__ == "__main__":
    main()
