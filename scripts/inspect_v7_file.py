"""Print the first rows, types, fills, and merged ranges of a 1C v7.7 workbook."""

from __future__ import annotations

import argparse
import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import xlrd
from rich.table import Table

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.parsers.v7_parser import _detect_excel_format
from src.utils.logger import console

ORDERS_DIR = PROJECT_ROOT / "data" / "orders"
PREVIEW_ROWS = 35
PREVIEW_COLS = 8
COL_LETTERS = "ABCDEFGH"


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect a 1C 7.7 .xls/.xlsx print form")
    parser.add_argument(
        "path",
        nargs="?",
        help="Path to workbook; defaults to the first file in data/orders/",
    )
    args = parser.parse_args()
    workbook_path = _resolve_path(args.path)
    payload = workbook_path.read_bytes()
    fmt = _detect_excel_format(payload, filename=workbook_path.name, path=workbook_path)

    console.rule(f"[bold]{workbook_path.name}[/bold] ({fmt})")
    if fmt == "xls":
        _inspect_xls(payload)
    else:
        _inspect_xlsx(payload)


def _resolve_path(explicit: str | None) -> Path:
    if explicit:
        candidate = Path(explicit)
        if not candidate.is_absolute():
            candidate = PROJECT_ROOT / candidate
        if not candidate.exists():
            raise SystemExit(f"File not found: {candidate}")
        return candidate

    if not ORDERS_DIR.exists():
        raise SystemExit(f"Orders directory not found: {ORDERS_DIR}")
    candidates = sorted(ORDERS_DIR.glob("*.xlsx")) + sorted(ORDERS_DIR.glob("*.xls"))
    if not candidates:
        raise SystemExit(f"No .xls/.xlsx files in {ORDERS_DIR}")
    return candidates[0]


def _inspect_xlsx(payload: bytes) -> None:
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    try:
        sheet = workbook.active
        table = _preview_table()
        max_row = min(sheet.max_row or 0, PREVIEW_ROWS)
        for row in range(1, max_row + 1):
            cells: list[str] = [str(row)]
            for col in range(1, PREVIEW_COLS + 1):
                cell = sheet.cell(row, col)
                cells.append(_format_preview_cell(cell.value, _openpyxl_type(cell.value), _openpyxl_fill(cell)))
            table.add_row(*cells)
        console.print(table)

        merged = [str(rng) for rng in sheet.merged_cells.ranges]
        _print_merged(merged)
    finally:
        workbook.close()


def _inspect_xls(payload: bytes) -> None:
    try:
        book = xlrd.open_workbook(file_contents=payload, formatting_info=True)
    except Exception:
        book = xlrd.open_workbook(file_contents=payload, formatting_info=False)
    sheet = book.sheet_by_index(0)
    table = _preview_table()
    max_row = min(sheet.nrows, PREVIEW_ROWS)
    for row in range(max_row):
        cells: list[str] = [str(row + 1)]
        for col in range(PREVIEW_COLS):
            if col >= sheet.ncols:
                cells.append("—")
                continue
            cell = sheet.cell(row, col)
            value = _normalize_xls_value(cell, book)
            cells.append(_format_preview_cell(value, _xls_type(cell), _xls_fill(book, cell)))
        table.add_row(*cells)
    console.print(table)

    merged: list[str] = []
    for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", []) or []:
        merged.append(f"{_a1(rlo, clo)}:{_a1(rhi - 1, chi - 1)}")
    _print_merged(merged)


def _preview_table() -> Table:
    table = Table(title=f"Первые {PREVIEW_ROWS} строк (A–H)", show_lines=False, expand=True)
    table.add_column("Строка", style="bold", justify="right")
    for letter in COL_LETTERS:
        table.add_column(letter, overflow="fold")
    return table


def _format_preview_cell(value: object, type_name: str, fill: str) -> str:
    if value is None or str(value).strip() == "":
        display = "∅"
    else:
        display = str(value).replace("\n", " ")
        if len(display) > 48:
            display = display[:45] + "…"
    fill_part = f" fill={fill}" if fill and fill not in {"00000000", "000000"} else ""
    return f"{display}\n[dim]{type_name}{fill_part}[/dim]"


def _openpyxl_type(value: object) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, datetime):
        return "date"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    return type(value).__name__


def _openpyxl_fill(cell) -> str:
    fill = cell.fill
    if fill is None or fill.start_color is None or fill.start_color.rgb is None:
        return ""
    rgb = str(fill.start_color.rgb)
    if rgb in {"00000000", "000000", "None"}:
        return ""
    return rgb


def _xls_type(cell: xlrd.sheet.Cell) -> str:
    mapping = {
        xlrd.XL_CELL_EMPTY: "empty",
        xlrd.XL_CELL_TEXT: "text",
        xlrd.XL_CELL_NUMBER: "number",
        xlrd.XL_CELL_DATE: "date",
        xlrd.XL_CELL_BOOLEAN: "bool",
        xlrd.XL_CELL_BLANK: "blank",
        xlrd.XL_CELL_ERROR: "error",
    }
    return mapping.get(cell.ctype, str(cell.ctype))


def _xls_fill(book: xlrd.Book, cell: xlrd.sheet.Cell) -> str:
    if not getattr(book, "formatting_info", False):
        return ""
    xf_index = cell.xf_index
    if xf_index is None or xf_index >= len(book.xf_list):
        return ""
    xf = book.xf_list[xf_index]
    rgb = book.colour_map.get(xf.background.pattern_colour_index)
    if rgb is None:
        return ""
    red, green, blue = rgb
    return f"FF{red:02X}{green:02X}{blue:02X}"


def _normalize_xls_value(cell: xlrd.sheet.Cell, book: xlrd.Book) -> object:
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        value = cell.value
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    return cell.value


def _print_merged(ranges: list[str]) -> None:
    console.print()
    if not ranges:
        console.print("[dim]Объединённые диапазоны: нет[/dim]")
        return
    console.print(f"[bold]Объединённые диапазоны ({len(ranges)}):[/bold]")
    for rng in ranges:
        console.print(f"  • {rng}")


def _a1(row0: int, col0: int) -> str:
    col = col0 + 1
    letters = ""
    while col:
        col, rem = divmod(col - 1, 26)
        letters = chr(65 + rem) + letters
    return f"{letters}{row0 + 1}"


if __name__ == "__main__":
    main()
