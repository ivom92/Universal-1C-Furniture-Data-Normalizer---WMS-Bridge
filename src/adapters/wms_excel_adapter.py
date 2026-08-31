"""WMS Excel export adapter — 5-column contract for warehouse import."""

from __future__ import annotations

import io
import json
import re
from datetime import date, datetime
from pathlib import Path
from collections.abc import Mapping
from typing import Any, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from src.models import MatchDecision
from src.utils.reporter import count_without_barcode

WMS_COLUMNS: list[str] = ["№", "Наименование", "Штрихкод", "Количество", "Заказчик"]
WMS_SHEET_NAME = "Импорт_WMS"
SUMMARY_SHEET_NAME = "Сводка_Отбора"
_TEXT_FORMAT = "@"
_QTY_FORMAT = "0"
_LINE_NUMBER_FORMAT = "0"
_MATCHED_STATUSES = frozenset({"MATCHED_AUTO", "MATCHED_LLM"})
_NO_BARCODE_METHODS = frozenset(
    {
        "AUTO_NO_BARCODE",
        "MATCHED_AUTO_NO_BARCODE",
        "LLM_NO_BARCODE",
        "exact_article_no_barcode",
    }
)
_EMPTY_BARCODE_TOKENS = frozenset({"none", "nan", "null"})
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F3542")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)
_BODY_ALIGNMENT = Alignment(vertical="center")
_CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center")
_THIN_SIDE = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)
_TOTALS_FILL = PatternFill(fill_type="solid", fgColor="F1F2F6")
_TOTALS_FONT = Font(bold=True, color="1E272E")
_TOTALS_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=Side(style="thin", color="2F3542"),
    bottom=Side(style="double", color="2F3542"),
)
_SUMMARY_TITLE_FONT = Font(bold=True, size=16, color="2F3542")
_SUMMARY_LABEL_FONT = Font(bold=True, color="2F3542")
_QUARANTINE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="F39C12")
_NO_BARCODE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="57606F")
_SUMMARY_ALT_FILL = PatternFill(fill_type="solid", fgColor="F8F9FA")
_MIN_COLUMN_WIDTHS = (9, 50, 18, 12, 25)
_WIDTH_PADDING = 3
_MAX_COLUMN_WIDTH = 80
_PREVIEW_ROWS = 20
_WMS_COL_COUNT = 5
_EAN13_RE = re.compile(r"^\d{13}$")
STATUS_OPERATOR_SCAN = "Введен вручную / со сканера"
STATUS_NO_FACTORY_BARCODE = "Заводской ШК отсутствует (фурнитура/погонаж)"
STATUS_QUARANTINE_PICK = "Нестандартный заказной размер (ручной отбор)"
NO_BARCODE_INSTRUCTION = (
    "Позиции ниже не имеют заводского штрихкода в каталоге 1С 8. "
    "Если штрихкод не был внесен через веб-интерфейс перед выгрузкой, "
    "внесите его вручную в WMS или отсканируйте при фактической приемке на склад."
)


class WMSExcelAdapter:
    """Generate a standardized 5-column Excel file for WMS warehouse import."""

    def export(
        self,
        decisions: list[MatchDecision],
        customer_name: str,
        output_path: Union[str, Path],
        *,
        source_name: str | None = None,
        processed_at: datetime | date | None = None,
        overrides: Mapping[int, str] | Mapping[str, str] | None = None,
    ) -> Path:
        """Write WMS Excel to disk and return the resolved output path."""
        destination = Path(output_path)
        destination.parent.mkdir(parents=True, exist_ok=True)

        workbook = self._build_workbook(
            decisions,
            customer_name,
            source_name=source_name,
            processed_at=processed_at,
            overrides=overrides,
        )
        workbook.save(destination)
        self.write_preview_sidecar(decisions, customer_name, destination)
        return destination.resolve()

    def export_to_bytes(
        self,
        decisions: list[MatchDecision],
        customer_name: str,
        *,
        source_name: str | None = None,
        processed_at: datetime | date | None = None,
        overrides: Mapping[int, str] | Mapping[str, str] | None = None,
    ) -> io.BytesIO:
        """Build WMS Excel in memory for Streamlit download buttons."""
        buffer = io.BytesIO()
        workbook = self._build_workbook(
            decisions,
            customer_name,
            source_name=source_name,
            processed_at=processed_at,
            overrides=overrides,
        )
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def build_download_filename(customer_name: str, date_str: str | None = None) -> str:
        """Return ``WMS_Импорт_{customer}_{date}.xlsx`` with safe filename characters."""
        safe_customer = re.sub(r'[<>:"/\\|?*]', "_", customer_name.strip()) or "Заказ"
        stamp = date_str or date.today().isoformat()
        return f"WMS_Импорт_{safe_customer}_{stamp}.xlsx"

    @classmethod
    def count_matched_with_barcode(
        cls,
        decisions: list[MatchDecision],
        overrides: Mapping[int, str] | Mapping[str, str] | None = None,
    ) -> int:
        return _count_matched_with_barcode(decisions, cls.normalize_overrides(overrides))

    @staticmethod
    def sort_decisions(decisions: list[MatchDecision]) -> list[MatchDecision]:
        items_sorted = sorted(decisions, key=lambda x: x.order_line_number)
        return items_sorted

    @staticmethod
    def wms_data_last_row(item_count: int) -> int:
        """Excel row of the last data line (header is row 1)."""
        return max(1, int(item_count) + 1)

    @staticmethod
    def wms_totals_row(item_count: int) -> int:
        """Excel row of the ИТОГО line (N+2 when N items occupy rows 2..N+1)."""
        return int(item_count) + 2

    @staticmethod
    def is_valid_ean13(value: object | None) -> bool:
        """True when value is exactly 13 ASCII digits (EAN-13, stored as str)."""
        if value is None:
            return False
        return bool(_EAN13_RE.fullmatch(str(value).strip()))

    @classmethod
    def normalize_overrides(
        cls,
        overrides: Mapping[int, str] | Mapping[str, str] | None,
    ) -> dict[int, str]:
        """Coerce operator barcode overrides to ``{line_number: ean_str}``."""
        if not overrides:
            return {}
        normalized: dict[int, str] = {}
        for key, value in overrides.items():
            barcode = cls._clean_barcode(value)
            if not barcode:
                continue
            normalized[int(key)] = barcode
        return normalized

    @classmethod
    def map_row(
        cls,
        decision: MatchDecision,
        customer_name: str,
        overrides: Mapping[int, str] | Mapping[str, str] | None = None,
    ) -> dict[str, Any]:
        """Map a match decision to the 5-column WMS contract (same as Streamlit export)."""
        nomenclature, barcode, quantity = cls._row_values(decision, overrides=overrides)
        return {
            "№": int(decision.order_line_number),
            "Наименование": nomenclature,
            "Штрихкод": barcode,
            "Количество": quantity,
            "Заказчик": decision.raw_block.customer_override or customer_name,
        }

    @classmethod
    def preview_rows(
        cls,
        decisions: list[MatchDecision],
        customer_name: str,
        limit: int = _PREVIEW_ROWS,
    ) -> list[dict[str, Any]]:
        ordered = cls.sort_decisions(decisions)
        return [cls.map_row(decision, customer_name) for decision in ordered[:limit]]

    @classmethod
    def write_preview_sidecar(
        cls,
        decisions: list[MatchDecision],
        customer_name: str,
        xlsx_path: Union[str, Path],
    ) -> Path:
        """Persist first-N mapped rows next to the workbook for export audit."""
        destination = Path(xlsx_path)
        sidecar = destination.parent / "last_wms_preview.json"
        ordered = cls.sort_decisions(decisions)
        payload = {
            "xlsx": destination.name,
            "customer_name": customer_name,
            "row_count": len(ordered),
            "qty_sum": sum(int(decision.raw_block.quantity) for decision in ordered),
            "preview": cls.preview_rows(ordered, customer_name),
        }
        sidecar.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return sidecar

    def _build_workbook(
        self,
        decisions: list[MatchDecision],
        customer_name: str,
        *,
        source_name: str | None = None,
        processed_at: datetime | date | None = None,
        overrides: Mapping[int, str] | Mapping[str, str] | None = None,
    ) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = WMS_SHEET_NAME

        items_sorted = self.sort_decisions(decisions)
        item_count = len(items_sorted)
        data_last_row = self.wms_data_last_row(item_count)
        resolved_overrides = self.normalize_overrides(overrides)

        self._write_header(worksheet)
        for row_index, decision in enumerate(items_sorted, start=2):
            nomenclature, barcode, quantity = self._row_values(
                decision,
                overrides=resolved_overrides,
            )
            self._write_data_row(
                worksheet,
                row_index=row_index,
                line_number=int(decision.order_line_number),
                nomenclature=nomenclature,
                barcode=barcode,
                quantity=quantity,
                customer_name=decision.raw_block.customer_override or customer_name,
            )

        self._style_used_range(worksheet, data_last_row)
        self._write_totals_row(worksheet, item_count)
        totals_row = self.wms_totals_row(item_count)
        self._autofit_columns(worksheet, totals_row)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:E{data_last_row}"

        self._write_summary_sheet(
            workbook,
            items_sorted,
            customer_name=customer_name,
            source_name=source_name,
            processed_at=processed_at,
            overrides=resolved_overrides,
        )
        return workbook

    @staticmethod
    def _write_header(worksheet: Worksheet) -> None:
        worksheet.row_dimensions[1].height = 26
        for col_index, header in enumerate(WMS_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

    @classmethod
    def _write_totals_row(cls, worksheet: Worksheet, item_count: int) -> None:
        totals_row = cls.wms_totals_row(item_count)
        data_last_row = cls.wms_data_last_row(item_count)
        formula = f"=SUM(D2:D{data_last_row})"

        marker = worksheet.cell(row=totals_row, column=1, value="Σ")
        name_cell = worksheet.cell(
            row=totals_row,
            column=2,
            value=f"ИТОГО (Позиций: {item_count})",
        )
        barcode_cell = worksheet.cell(row=totals_row, column=3, value=None)
        qty_cell = worksheet.cell(row=totals_row, column=4, value=formula)
        customer_cell = worksheet.cell(row=totals_row, column=5, value=None)

        qty_cell.number_format = _QTY_FORMAT
        for cell in (marker, name_cell, barcode_cell, qty_cell, customer_cell):
            cell.font = _TOTALS_FONT
            cell.fill = _TOTALS_FILL
            cell.border = _TOTALS_BORDER
        marker.alignment = _CENTER_ALIGNMENT
        name_cell.alignment = _BODY_ALIGNMENT
        qty_cell.alignment = _RIGHT_ALIGNMENT

    @classmethod
    def _row_values(
        cls,
        decision: MatchDecision,
        overrides: Mapping[int, str] | Mapping[str, str] | None = None,
    ) -> tuple[str, str, int]:
        block = decision.raw_block
        quantity = int(block.quantity)
        override = cls.normalize_overrides(overrides).get(int(decision.order_line_number), "")

        if decision.status != "QUARANTINE" and decision.matched_entity is not None:
            entity = decision.matched_entity
            barcode = override or cls._clean_barcode(entity.barcode)
            return entity.nomenclature, barcode, quantity

        name = (block.client_description or "").strip()
        return name, override, quantity

    @staticmethod
    def _clean_barcode(value: object | None) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if not text or text.lower() in _EMPTY_BARCODE_TOKENS:
            return ""
        return text

    @staticmethod
    def _write_data_row(
        worksheet: Worksheet,
        *,
        row_index: int,
        line_number: int,
        nomenclature: str,
        barcode: str,
        quantity: int,
        customer_name: str,
    ) -> None:
        number_cell = worksheet.cell(row=row_index, column=1, value=int(line_number))
        number_cell.number_format = _LINE_NUMBER_FORMAT
        number_cell.alignment = _CENTER_ALIGNMENT

        name_cell = worksheet.cell(row=row_index, column=2, value=nomenclature)
        name_cell.alignment = _BODY_ALIGNMENT

        barcode_cell = worksheet.cell(row=row_index, column=3)
        barcode_cell.number_format = _TEXT_FORMAT
        barcode_cell.data_type = "s"
        barcode_cell.alignment = _BODY_ALIGNMENT
        if barcode:
            barcode_cell.value = str(barcode)
        else:
            barcode_cell.value = None

        qty_cell = worksheet.cell(row=row_index, column=4, value=int(quantity))
        qty_cell.number_format = _QTY_FORMAT
        qty_cell.alignment = _CENTER_ALIGNMENT

        customer_cell = worksheet.cell(row=row_index, column=5, value=customer_name)
        customer_cell.alignment = _BODY_ALIGNMENT

    @staticmethod
    def _style_used_range(worksheet: Worksheet, last_row: int) -> None:
        for row_index in range(1, last_row + 1):
            for col_index in range(1, _WMS_COL_COUNT + 1):
                worksheet.cell(row=row_index, column=col_index).border = _THIN_BORDER

    @staticmethod
    def _autofit_columns(worksheet: Worksheet, last_row: int) -> None:
        for col_index, min_width in enumerate(_MIN_COLUMN_WIDTHS, start=1):
            max_len = 0
            for row_index in range(1, last_row + 1):
                value = worksheet.cell(row=row_index, column=col_index).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            width = max(min_width, max_len + _WIDTH_PADDING)
            width = min(width, _MAX_COLUMN_WIDTH)
            if col_index == 1:
                width = min(max(width, 8), 10)
            worksheet.column_dimensions[get_column_letter(col_index)].width = width

    def _write_summary_sheet(
        self,
        workbook: Workbook,
        decisions: list[MatchDecision],
        *,
        customer_name: str,
        source_name: str | None,
        processed_at: datetime | date | None,
        overrides: Mapping[int, str] | Mapping[str, str] | None = None,
    ) -> None:
        worksheet = workbook.create_sheet(title=SUMMARY_SHEET_NAME)
        stamp = _format_processed_at(processed_at)
        resolved_overrides = self.normalize_overrides(overrides)
        item_count = len(decisions)
        qty_sum = sum(int(decision.raw_block.quantity) for decision in decisions)
        matched_with_barcode = _count_matched_with_barcode(decisions, resolved_overrides)
        without_barcode = count_without_barcode(decisions, resolved_overrides)
        quarantine = [
            decision
            for decision in decisions
            if decision.status == "QUARANTINE"
            and int(decision.order_line_number) not in resolved_overrides
        ]
        no_barcode_rows = [
            decision
            for decision in decisions
            if _is_no_barcode_position(decision)
            or (
                decision.status == "QUARANTINE"
                and int(decision.order_line_number) in resolved_overrides
            )
        ]

        worksheet.merge_cells("A1:D1")
        title = worksheet.cell(row=1, column=1, value="Сводка отбора для склада")
        title.font = _SUMMARY_TITLE_FONT
        worksheet.row_dimensions[1].height = 24

        card_rows = (
            (3, "Заказчик", customer_name or "—"),
            (4, "Источник", source_name or "—"),
            (5, "Дата обработки", stamp),
            (6, "Всего строк", item_count),
            (7, "Всего упаковок/мест", qty_sum),
            (8, "Сопоставлено с ШК", matched_with_barcode),
            (9, "Без ШК", without_barcode),
            (10, "Карантин", len(quarantine)),
        )
        for row_index, label, value in card_rows:
            label_cell = worksheet.cell(row=row_index, column=1, value=label)
            label_cell.font = _SUMMARY_LABEL_FONT
            value_cell = worksheet.cell(row=row_index, column=2, value=value)
            value_cell.alignment = _BODY_ALIGNMENT

        next_row = 12
        next_row = self._write_operator_table(
            worksheet,
            start_row=next_row,
            title="Таблица 1. Позиции карантина (ручной отбор)",
            headers=("№", "Наименование", "Кол-во", "Причина"),
            header_fill=_QUARANTINE_HEADER_FILL,
            rows=[
                (
                    int(decision.order_line_number),
                    self._row_values(decision, overrides=resolved_overrides)[0],
                    int(decision.raw_block.quantity),
                    _quarantine_reason(decision),
                )
                for decision in quarantine
            ],
            empty_message="Карантинных позиций нет.",
        )
        next_row += 2
        worksheet.merge_cells(
            start_row=next_row,
            start_column=1,
            end_row=next_row,
            end_column=4,
        )
        instruction = worksheet.cell(row=next_row, column=1, value=NO_BARCODE_INSTRUCTION)
        instruction.alignment = Alignment(wrap_text=True, vertical="center")
        instruction.font = Font(italic=True, color="2F3542", size=10)
        worksheet.row_dimensions[next_row].height = 48
        next_row += 2
        self._write_operator_table(
            worksheet,
            start_row=next_row,
            title="Таблица 2. Позиции без ШК (фурнитура / погонаж)",
            headers=("№", "Наименование", "Кол-во", "Способ"),
            header_fill=_NO_BARCODE_HEADER_FILL,
            rows=[
                (
                    int(decision.order_line_number),
                    self._row_values(decision, overrides=resolved_overrides)[0],
                    int(decision.raw_block.quantity),
                    _localized_summary_method(decision, resolved_overrides),
                )
                for decision in no_barcode_rows
            ],
            empty_message="Позиций без штрихкода нет.",
        )

        worksheet.column_dimensions["A"].width = 22
        worksheet.column_dimensions["B"].width = 70
        worksheet.column_dimensions["C"].width = 12
        worksheet.column_dimensions["D"].width = 42
        worksheet.freeze_panes = "A3"
        worksheet.page_setup.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.orientation = "landscape"
        worksheet.print_title_rows = "1:1"

    @staticmethod
    def _write_operator_table(
        worksheet: Worksheet,
        *,
        start_row: int,
        title: str,
        headers: tuple[str, ...],
        header_fill: PatternFill,
        rows: list[tuple[Any, ...]],
        empty_message: str,
    ) -> int:
        title_cell = worksheet.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=12, color="2F3542")
        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=len(headers),
        )

        header_row = start_row + 1
        for col_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=col_index, value=header)
            cell.font = _HEADER_FONT
            cell.fill = header_fill
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

        if not rows:
            empty_row = header_row + 1
            worksheet.merge_cells(
                start_row=empty_row,
                start_column=1,
                end_row=empty_row,
                end_column=len(headers),
            )
            empty_cell = worksheet.cell(row=empty_row, column=1, value=empty_message)
            empty_cell.font = Font(italic=True, color="57606F")
            return empty_row

        current = header_row
        for offset, values in enumerate(rows, start=1):
            current = header_row + offset
            for col_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=current, column=col_index, value=value)
                cell.border = _THIN_BORDER
                cell.alignment = _CENTER_ALIGNMENT if col_index in {1, 3} else _BODY_ALIGNMENT
                if offset % 2 == 0:
                    cell.fill = _SUMMARY_ALT_FILL
        return current


def auto_filter_data_last_row(worksheet: Worksheet) -> int:
    """Last data row covered by the WMS autofilter (excludes ИТОГО)."""
    ref = worksheet.auto_filter.ref
    if not ref:
        raise ValueError("WMS sheet has no auto_filter.ref")
    _min_col, _min_row, _max_col, max_row = range_boundaries(ref)
    return int(max_row)


def _format_processed_at(value: datetime | date | None) -> str:
    if value is None:
        return date.today().isoformat()
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return value.isoformat()


def _count_matched_with_barcode(
    decisions: list[MatchDecision],
    overrides: Mapping[int, str] | None = None,
) -> int:
    resolved = WMSExcelAdapter.normalize_overrides(overrides)
    total = 0
    for decision in decisions:
        line = int(decision.order_line_number)
        if line in resolved:
            total += 1
            continue
        if decision.status not in _MATCHED_STATUSES:
            continue
        barcode = decision.matched_entity.barcode if decision.matched_entity is not None else None
        if WMSExcelAdapter._clean_barcode(barcode):
            total += 1
    return total


def _is_no_barcode_position(decision: MatchDecision) -> bool:
    if decision.status == "QUARANTINE":
        return False
    method = (decision.match_method or "").strip()
    if method in _NO_BARCODE_METHODS:
        return True
    if decision.status not in _MATCHED_STATUSES:
        return False
    barcode = decision.matched_entity.barcode if decision.matched_entity is not None else None
    return not WMSExcelAdapter._clean_barcode(barcode)


def _localized_summary_method(
    decision: MatchDecision,
    overrides: Mapping[int, str] | None = None,
) -> str:
    resolved = WMSExcelAdapter.normalize_overrides(overrides)
    if int(decision.order_line_number) in resolved:
        return STATUS_OPERATOR_SCAN
    if decision.status == "QUARANTINE":
        return STATUS_QUARANTINE_PICK
    return STATUS_NO_FACTORY_BARCODE


def _quarantine_reason(decision: MatchDecision) -> str:
    detail = (decision.status_detail or "").strip()
    if detail:
        if detail == "Нестандартный заказной размер":
            return STATUS_QUARANTINE_PICK
        return detail
    return STATUS_QUARANTINE_PICK
