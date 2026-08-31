"""State-machine parser for 1C v7.7 visual picking-list workbooks."""

from __future__ import annotations

import io
import logging
import re
import unicodedata
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterator, Optional, Protocol, Union

import openpyxl
import pandas as pd
import xlrd
from openpyxl.worksheet.worksheet import Worksheet
from xlrd.sheet import Sheet as XlrdSheet

from src.models import RawOrderBlock, V7ParseResult
from src.preprocessor.normalizer import heal_mojibake

logger = logging.getLogger(__name__)

_CYRILLIC_RE = re.compile(r"[А-Яа-яЁё]")
_MOJIBAKE_LATIN_RE = re.compile(r"[\u00C0-\u00FF]")
_HTML_ENCODINGS = ("utf-8", "cp1251", "cp866")

HEADER_LABEL_PATTERN = re.compile(
    r"(рег\.?\s*склад|региональный\s*склад|"
    r"склад[- ]получатель|подразделение[- ]получатель|грузополучатель|"
    r"покупатель|получатель|контрагент|заказчик|клиент|куда|кому)\s*:?\s*(.*)",
    re.IGNORECASE,
)
DOCUMENT_TITLE_PATTERN = re.compile(
    r"(перемещение|отборочн\w*(?:\s+лист)?|накладн\w*|расходн\w*)\b.{3,}",
    re.IGNORECASE | re.DOTALL,
)
CELL_LOCATION_PATTERN = re.compile(r"^Р\d+[\.\w\d]+\s+", re.IGNORECASE)
CELL_LOCATION_TOKEN_PATTERN = re.compile(r"^Р\d+[\.\w\d]+$", re.IGNORECASE)
WAREHOUSE_RACK_TOKEN_RE = re.compile(r"Р\d+[\.\w\d]+", re.IGNORECASE)
WAREHOUSE_SECTION_TOKEN_RE = re.compile(r"секция\s+\d+", re.IGNORECASE)
LOCATION_HEADER_PATTERN = re.compile(
    r"линейка|секция|ячейка|адрес|(?:^|[\s/])место(?:[\s/]|$)",
    re.IGNORECASE,
)
QUANTITY_HEADER_PATTERN = re.compile(
    r"кол-во|количество|к-во|^кол\.?$",
    re.IGNORECASE,
)
UNIT_HEADER_PATTERN = re.compile(r"^ед\.?$", re.IGNORECASE)
TYPE_HEADER_PATTERN = re.compile(r"^тип$", re.IGNORECASE)
NOTE_HEADER_PATTERN = re.compile(r"отметк", re.IGNORECASE)
CODE_HEADER_PATTERN = re.compile(r"^код$", re.IGNORECASE)
NUM_HEADER_PATTERN = re.compile(
    r"№|п/п|^номер$|^n$|^no\.?$|^код$",
    re.IGNORECASE,
)
TABLE_NAME_TOKENS = ("товар", "номенклатура", "наименование", "упаковка")
TABLE_INDEX_TOKENS = ("№", "код", "кол-во", "количество", "к-во")

CUSTOMER_LABELS = ("Покупатель:", "Получатель:", "Контрагент:")
CUSTOMER_LABEL = CUSTOMER_LABELS[0]
MAIN_ROW_FILL = "FFE0FFE0"
ALIAS_ROW_FILL = "FFFFFFC0"
DESCRIPTION_START_COL = 2
DESCRIPTION_END_COL = 6
SERVICE_TEXT_COLS = (2, 3, 4, 5, 6)
QUANTITY_COLS = (8, 5, 4, 6, 3)
HEADER_SCAN_ROWS = 40
HEADER_SCAN_COLS = 12
DOCUMENT_TITLE_ROWS = 10
TABLE_HEADER_SCAN_ROWS = 50
HTML_PEEK_BYTES = 200

OLE2_MAGIC = b"\xd0\xcf\x11\xe0"
ZIP_MAGIC = b"PK\x03\x04"
HTML_PREFIXES = (b"<html", b"<!doctype", b"<table", b"<?xml")
FOOTER_ROW_PATTERN = re.compile(r"итого\s+мест|экспедитор|кладовщик", re.IGNORECASE)

V7Source = Union[str, Path, bytes, io.BytesIO]


@dataclass(frozen=True)
class TableLayout:
    """Dynamic spatial column map for a 1C v7.7 printed table."""

    col_num: int = 1
    col_name: int = 2
    col_type: int | None = 7
    col_qty: int | None = 8
    col_rack: int | None = None
    desc_cols: tuple[int, ...] = (2, 3, 4, 5, 6)
    skip_desc: frozenset[int] = frozenset()


DEFAULT_LAYOUT = TableLayout()


class _SheetView(Protocol):
    max_row: int
    max_column: int

    def cell_value(self, row: int, col: int) -> object: ...

    def cell_fill_rgb(self, row: int, col: int) -> str: ...


class _OpenpyxlSheetView:
    def __init__(self, worksheet: Worksheet) -> None:
        self._worksheet = worksheet
        self.max_row = int(worksheet.max_row or 0)
        self.max_column = int(worksheet.max_column or 0)

    def cell_value(self, row: int, col: int) -> object:
        return self._worksheet.cell(row, col).value

    def cell_fill_rgb(self, row: int, col: int) -> str:
        fill = self._worksheet.cell(row, col).fill
        if fill is None or fill.start_color is None or fill.start_color.rgb is None:
            return "00000000"
        return str(fill.start_color.rgb)


class _XlrdSheetView:
    def __init__(self, book: xlrd.Book, sheet: XlrdSheet) -> None:
        self._book = book
        self._sheet = sheet
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def cell_value(self, row: int, col: int) -> object:
        if row < 1 or col < 1 or row > self._sheet.nrows or col > self._sheet.ncols:
            return None
        cell = self._sheet.cell(row - 1, col - 1)
        return _normalize_xlrd_cell(cell, self._book)

    def cell_fill_rgb(self, row: int, col: int) -> str:
        if not getattr(self._book, "formatting_info", False):
            return "00000000"
        if row < 1 or col < 1 or row > self._sheet.nrows or col > self._sheet.ncols:
            return "00000000"
        cell = self._sheet.cell(row - 1, col - 1)
        xf_index = cell.xf_index
        if xf_index is None or xf_index >= len(self._book.xf_list):
            return "00000000"
        xf = self._book.xf_list[xf_index]
        colour_index = xf.background.pattern_colour_index
        rgb = self._book.colour_map.get(colour_index)
        if rgb is None:
            return "00000000"
        red, green, blue = rgb
        return f"FF{red:02X}{green:02X}{blue:02X}"


class _MatrixSheetView:
    """In-memory grid used for HTML tables saved with an .xls extension."""

    def __init__(self, rows: list[list[object]]) -> None:
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(row) for row in rows), default=0)

    def cell_value(self, row: int, col: int) -> object:
        if row < 1 or col < 1 or row > self.max_row or col > len(self._rows[row - 1]):
            return None
        value = self._rows[row - 1][col - 1]
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = normalize_incoming_text(value)
            return cleaned or None
        return value

    def cell_fill_rgb(self, row: int, col: int) -> str:
        return "00000000"


_NUMERO_GUARD = "\ue000"


def normalize_incoming_text(value: object) -> str:
    """NFKC-normalize workbook text, keep '№', and collapse non-breaking spaces."""
    if value is None:
        return ""
    text = heal_mojibake(str(value))
    text = text.replace("№", _NUMERO_GUARD)
    text = unicodedata.normalize("NFKC", text)
    text = text.replace(_NUMERO_GUARD, "№")
    text = text.replace("\xa0", " ").replace("\u202f", " ").replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def strip_cell_location_prefix(raw_name: str) -> str:
    """Remove warehouse topology tokens (Р1.16.Я2, Секция 12) from product text."""
    return sanitize_warehouse_topology(raw_name)


def sanitize_warehouse_topology(raw_name: str) -> str:
    """Strip rack/section tokens whether they sit in their own cell or are glued into names."""
    cleaned = str(raw_name).strip()
    if not cleaned:
        return ""
    if CELL_LOCATION_TOKEN_PATTERN.fullmatch(cleaned):
        return ""
    cleaned = WAREHOUSE_RACK_TOKEN_RE.sub(" ", cleaned)
    cleaned = WAREHOUSE_SECTION_TOKEN_RE.sub(" ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def parse_v7_order(
    source: V7Source,
    filename: str | None = None,
) -> V7ParseResult:
    """Parse a 1C v7.7 order file (.xlsx or .xls) into customer metadata and 3-row blocks."""
    workbook_path, payload = _resolve_source(source)
    fmt = _detect_excel_format(payload, filename=filename, path=workbook_path)

    source_name = filename or (workbook_path.name if workbook_path is not None else "unknown")

    if fmt == "html":
        return _parse_sheet(_load_html_sheet(workbook_path, payload), source_name)

    if fmt == "xls":
        sheet = _load_xls_sheet(workbook_path, payload)
        return _parse_sheet(sheet, source_name)

    workbook = _open_xlsx_workbook(workbook_path, payload)
    try:
        return _parse_sheet(_OpenpyxlSheetView(workbook.active), source_name)
    finally:
        workbook.close()


@contextmanager
def open_order_sheet(
    source: V7Source,
    filename: str | None = None,
) -> Iterator[_SheetView]:
    """Yield the first worksheet of a v7.7 workbook (.xlsx / .xls / HTML-as-xls)."""
    workbook_path, payload = _resolve_source(source)
    fmt = _detect_excel_format(payload, filename=filename, path=workbook_path)
    workbook = None
    try:
        if fmt == "html":
            yield _load_html_sheet(workbook_path, payload)
        elif fmt == "xls":
            yield _load_xls_sheet(workbook_path, payload)
        else:
            workbook = _open_xlsx_workbook(workbook_path, payload)
            yield _OpenpyxlSheetView(workbook.active)
    finally:
        if workbook is not None:
            workbook.close()


def _parse_sheet(sheet: _SheetView, source_name: str) -> V7ParseResult:
    customer_name = _extract_customer_name(sheet, source_name)
    blocks = _parse_blocks_state_machine(sheet)
    return V7ParseResult(customer_name=customer_name, blocks=blocks)


def parse_v7_blocks_from_sheet(sheet: _SheetView) -> list[RawOrderBlock]:
    """Parse 3-row item blocks directly from an already-open (possibly windowed) sheet.

    Used by the document splitter to run the standard-picking-list cascade
    against a single section of a combined (``COMPOSITE_PICKING_LIST``) workbook.
    """
    return _parse_blocks_state_machine(sheet)


def extract_customer_name_from_sheet(sheet: _SheetView, source_name: str) -> str:
    """Extract the counterparty/customer name directly from an already-open sheet."""
    return _extract_customer_name(sheet, source_name)


def _resolve_source(source: V7Source) -> tuple[Path | None, bytes | None]:
    if isinstance(source, (bytes, bytearray, memoryview)):
        return None, bytes(source)
    if isinstance(source, io.BytesIO):
        position = source.tell()
        source.seek(0)
        payload = source.read()
        source.seek(position)
        return None, payload
    workbook_path = Path(source)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Order file not found: {workbook_path}")
    return workbook_path, None


def _detect_excel_format(
    payload: bytes | None,
    filename: str | None = None,
    path: Path | None = None,
) -> str:
    header = _peek_bytes(payload, path, HTML_PEEK_BYTES)
    if _looks_like_html(header):
        return "html"
    if header.startswith(OLE2_MAGIC):
        return "xls"
    if header.startswith(ZIP_MAGIC):
        return "xlsx"

    name = filename or (path.name if path is not None else "")
    suffix = Path(name).suffix.lower()
    if suffix == ".xls":
        return "xls"
    if suffix == ".xlsx":
        return "xlsx"

    raise ValueError(
        "Cannot detect Excel format: expected .xlsx (ZIP), .xls (OLE2/BIFF8), or HTML table"
    )


def _peek_bytes(payload: bytes | None, path: Path | None, size: int) -> bytes:
    if payload:
        return payload[:size]
    if path is not None:
        with path.open("rb") as handle:
            return handle.read(size)
    return b""


def _looks_like_html(header: bytes) -> bool:
    sample = header.lstrip(b"\xef\xbb\xbf \t\r\n")
    lowered = sample[:HTML_PEEK_BYTES].lower()
    return any(lowered.startswith(prefix) for prefix in HTML_PREFIXES)


def _open_xlsx_workbook(path: Path | None, payload: bytes | None):
    if payload is not None:
        return openpyxl.load_workbook(io.BytesIO(payload), read_only=False, data_only=True)
    assert path is not None
    return openpyxl.load_workbook(path, read_only=False, data_only=True)


def _load_xls_sheet(path: Path | None, payload: bytes | None) -> _SheetView:
    kwargs: dict[str, object] = {"formatting_info": True}
    if payload is not None:
        kwargs["file_contents"] = payload
    else:
        assert path is not None
        kwargs["filename"] = str(path)

    try:
        book = xlrd.open_workbook(**kwargs)
    except Exception:
        kwargs["formatting_info"] = False
        book = xlrd.open_workbook(**kwargs)

    if book.nsheets < 1:
        raise ValueError("Workbook contains no worksheets")
    return _XlrdSheetView(book, book.sheet_by_index(0))


def _looks_like_cp1251_mojibake(text: str) -> bool:
    """True when UTF-8/Latin-1 decoding produced fake Latin instead of Cyrillic."""
    if _CYRILLIC_RE.search(text):
        return False
    return bool(_MOJIBAKE_LATIN_RE.search(text))


def _decode_html_bytes(raw: bytes) -> str:
    """Auto-detect Windows-1251 vs UTF-8 for 1C v7.7 HTML tables saved as .xls."""
    payload = raw.lstrip(b"\xef\xbb\xbf")
    utf8_text: str | None = None
    try:
        utf8_text = payload.decode("utf-8")
    except UnicodeDecodeError:
        utf8_text = None

    if utf8_text is not None and not _looks_like_cp1251_mojibake(utf8_text):
        return utf8_text

    for encoding in _HTML_ENCODINGS[1:]:
        try:
            decoded = payload.decode(encoding)
        except UnicodeDecodeError:
            continue
        if _CYRILLIC_RE.search(decoded) or encoding == "cp1251":
            return decoded

    if utf8_text is not None:
        return utf8_text
    return payload.decode("cp1251", errors="replace")


def _load_html_sheet(path: Path | None, payload: bytes | None) -> _MatrixSheetView:
    raw = payload if payload is not None else path.read_bytes() if path is not None else b""
    text = _decode_html_bytes(raw)
    frames = _read_html_frames(text)
    if not frames:
        raise ValueError("HTML workbook contains no tables")
    frame = max(frames, key=lambda table: int(table.shape[0] * max(table.shape[1], 1)))
    rows: list[list[object]] = []
    for record in frame.fillna("").astype(object).values.tolist():
        cleaned: list[object] = []
        for cell in record:
            if cell is None:
                cleaned.append(None)
                continue
            text_cell = normalize_incoming_text(cell)
            if not text_cell or text_cell.lower() in {"nan", "none"}:
                cleaned.append(None)
            else:
                cleaned.append(text_cell)
        rows.append(cleaned)
    if not rows:
        raise ValueError("HTML workbook table is empty")
    return _MatrixSheetView(rows)


def _read_html_frames(text: str) -> list[pd.DataFrame]:
    try:
        frames = pd.read_html(io.StringIO(text), header=None, flavor="lxml")
        if frames:
            return list(frames)
    except Exception:
        logger.debug("pandas.read_html failed; falling back to BeautifulSoup", exc_info=True)
    try:
        from bs4 import BeautifulSoup
    except ImportError as exc:
        raise ValueError("HTML table parsing requires pandas/lxml or beautifulsoup4") from exc

    soup = BeautifulSoup(text, "html.parser")
    tables = soup.find_all("table")
    frames: list[pd.DataFrame] = []
    for table in tables:
        grid: list[list[str]] = []
        for tr in table.find_all("tr"):
            cells = [normalize_incoming_text(cell.get_text(" ", strip=True)) for cell in tr.find_all(["td", "th"])]
            if cells:
                grid.append(cells)
        if grid:
            width = max(len(row) for row in grid)
            padded = [row + [""] * (width - len(row)) for row in grid]
            frames.append(pd.DataFrame(padded))
    return frames


def _normalize_xlrd_cell(cell: xlrd.sheet.Cell, book: xlrd.Book) -> object:
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            parts = xlrd.xldate_as_tuple(cell.value, book.datemode)
            return datetime(*parts)
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        value = cell.value
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    text = str(cell.value)
    cleaned = normalize_incoming_text(text)
    return cleaned or None


def _extract_customer_name(worksheet: _SheetView, source_name: str) -> str:
    max_row = min(worksheet.max_row, HEADER_SCAN_ROWS)
    max_col = min(max(worksheet.max_column, 8), HEADER_SCAN_COLS)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            text = _cell_text(worksheet, row, col)
            if not text:
                continue
            match = HEADER_LABEL_PATTERN.search(text)
            if not match:
                continue
            inline = _clean_extracted_value(match.group(2))
            if _is_usable_party_name(inline):
                return inline
            from_neighbors = _scan_horizontal_value(worksheet, row, col, max_col)
            if from_neighbors:
                return from_neighbors
            below = _clean_extracted_value(_cell_text(worksheet, row + 1, col))
            if _is_usable_party_name(below):
                return below

    title = _extract_document_title(worksheet)
    if title:
        logger.warning(
            "Counterparty label not found in %s; using document title as customer_name",
            source_name,
        )
        return title

    fallback = _fallback_customer_name(source_name)
    logger.warning(
        "Counterparty label and document title not found in %s; using fallback %r",
        source_name,
        fallback,
    )
    return fallback


def _scan_horizontal_value(
    worksheet: _SheetView,
    row: int,
    col: int,
    max_col: int,
) -> str:
    for next_col in range(col + 1, max_col + 1):
        candidate = _clean_extracted_value(_cell_text(worksheet, row, next_col))
        if _is_usable_party_name(candidate):
            return candidate
    return ""


def _extract_document_title(worksheet: _SheetView) -> str:
    max_row = min(worksheet.max_row, DOCUMENT_TITLE_ROWS)
    max_col = min(max(worksheet.max_column, 8), HEADER_SCAN_COLS)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            text = _clean_extracted_value(_cell_text(worksheet, row, col))
            if not text:
                continue
            if DOCUMENT_TITLE_PATTERN.search(text):
                return text
    return ""


def _fallback_customer_name(source_name: str) -> str:
    filename = Path(source_name).name if source_name and source_name != "unknown" else ""
    if filename:
        return f"Перемещение ({filename})"
    return "Не указан (Перемещение)"


def _clean_extracted_value(text: str) -> str:
    cleaned = str(text).strip()
    cleaned = cleaned.strip(" \t\"'«»„“”'`")
    cleaned = re.sub(r"^[-–—]+", "", cleaned)
    cleaned = re.sub(r"[-–—]+$", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" \t:;-–—")
    return cleaned


def _is_usable_party_name(value: str) -> bool:
    if len(value) < 2:
        return False
    if HEADER_LABEL_PATTERN.fullmatch(value):
        return False
    label_only = HEADER_LABEL_PATTERN.match(value)
    if label_only and not _clean_extracted_value(label_only.group(2)):
        return False
    return True


def _parse_blocks_state_machine(worksheet: _SheetView) -> list[RawOrderBlock]:
    blocks: list[RawOrderBlock] = []
    pending_main: Optional[dict[str, object]] = None
    pending_alias: Optional[str] = None
    pending_service: str = ""
    pending_extras: list[str] = []
    layout = DEFAULT_LAYOUT
    header_row = _find_table_header_row(worksheet)
    table_started = header_row is not None
    if header_row is not None:
        layout = _detect_table_layout(worksheet, header_row)

    def flush_block() -> None:
        nonlocal pending_main, pending_alias, pending_service, pending_extras
        assert pending_main is not None
        factory_alias = pending_alias
        extras = [part for part in pending_extras if part]
        if extras:
            extra_text = " ".join(extras)
            factory_alias = f"{factory_alias} {extra_text}".strip() if factory_alias else extra_text
        if factory_alias:
            factory_alias = _strip_footer_noise(factory_alias) or None
        if factory_alias == "":
            factory_alias = None
        blocks.append(
            RawOrderBlock(
                line_number=int(pending_main["line_number"]),
                client_description=str(pending_main["client_description"]),
                item_type=str(pending_main["item_type"]),
                quantity=int(pending_main["quantity"]),
                factory_alias=factory_alias,
                order_service_line=pending_service or "",
                excel_row_start=int(pending_main["excel_row_start"]),
            )
        )
        pending_main = None
        pending_alias = None
        pending_service = ""
        pending_extras = []

    def start_main(row: int, line_number: int) -> None:
        nonlocal pending_main, pending_alias, pending_service, pending_extras
        pending_main = _read_main_payload(worksheet, row, layout, line_number=line_number)
        pending_alias = None
        pending_service = ""
        pending_extras = []

    for row in range(1, worksheet.max_row + 1):
        if _is_table_header_row(worksheet, row):
            table_started = True
            layout = _detect_table_layout(worksheet, row)
            if pending_main is not None:
                flush_block()
            continue

        if _is_document_footer_row(worksheet, row):
            if pending_main is not None:
                flush_block()
            break

        anchor = _integer_anchor(worksheet, row, layout)
        is_main = _is_main_row(
            worksheet,
            row,
            layout,
            require_fill=not table_started,
            line_number=anchor,
        )

        if is_main:
            if pending_main is not None:
                flush_block()
            assert anchor is not None
            start_main(row, anchor)
            continue

        if pending_main is None:
            continue

        if _is_alias_row(worksheet, row, layout) and not pending_alias:
            pending_alias = _row_name_text(worksheet, row, layout) or None
            continue

        service_line = _extract_service_line(worksheet, row, layout)
        if service_line:
            pending_service = f"{pending_service} {service_line}".strip()
            continue

        if _row_is_blank(worksheet, row):
            continue

        extra = _merge_description(worksheet, row, layout)
        if extra:
            pending_extras.append(extra)

    if pending_main is not None:
        flush_block()

    return blocks


def _read_main_payload(
    worksheet: _SheetView,
    row: int,
    layout: TableLayout,
    *,
    line_number: int | None = None,
) -> dict[str, object]:
    type_col = layout.col_type if layout.col_type is not None else 7
    parsed_line = line_number if line_number is not None else _integer_anchor(worksheet, row, layout)
    if parsed_line is None:
        raise ValueError(f"Line number is missing in main row {row}")
    quantity = _try_read_row_quantity(worksheet, row, layout)
    if quantity is None:
        quantity = 1
    return {
        "line_number": parsed_line,
        "client_description": _merge_description(worksheet, row, layout),
        "item_type": _cell_text(worksheet, row, type_col) if type_col <= worksheet.max_column else "",
        "quantity": quantity,
        "excel_row_start": row,
    }


def _find_table_header_row(worksheet: _SheetView) -> int | None:
    max_row = min(worksheet.max_row, TABLE_HEADER_SCAN_ROWS)
    for row in range(1, max_row + 1):
        if _is_table_header_row(worksheet, row):
            return row
    return None


def _strip_footer_noise(text: str) -> str:
    cleaned = FOOTER_ROW_PATTERN.split(text, maxsplit=1)[0].strip()
    return cleaned


def _is_document_footer_row(worksheet: _SheetView, row: int) -> bool:
    joined = " ".join(
        _cell_text(worksheet, row, col)
        for col in range(1, min(worksheet.max_column, 12) + 1)
    )
    return bool(joined) and bool(FOOTER_ROW_PATTERN.search(joined))


def _is_table_header_row(worksheet: _SheetView, row: int) -> bool:
    joined = " ".join(
        _cell_text(worksheet, row, col)
        for col in range(1, min(worksheet.max_column, 12) + 1)
    )
    if not joined:
        return False
    lowered = joined.lower()
    has_name = any(token in lowered for token in TABLE_NAME_TOKENS)
    has_location = bool(LOCATION_HEADER_PATTERN.search(joined))
    has_index = "№" in joined or any(token in lowered for token in ("код", "кол-во", "количество", "ед."))
    first_numbered = any(
        _cell_text(worksheet, row, col).isdigit()
        for col in range(1, min(worksheet.max_column, 4) + 1)
    )
    return (has_name or has_location) and has_index and not first_numbered


def _detect_table_layout(worksheet: _SheetView, header_row: int) -> TableLayout:
    max_col = min(max(worksheet.max_column, 8), 12)
    col_num: int | None = None
    col_rack: int | None = None
    col_name: int | None = None
    col_type: int | None = None
    col_qty: int | None = None
    skip: set[int] = set()

    for col in range(1, max_col + 1):
        header = _cell_text(worksheet, header_row, col)
        if not header:
            continue
        lowered = header.lower().strip()
        if "№" in header or "п/п" in lowered or lowered in {"номер", "n", "no", "no."}:
            col_num = col
            skip.add(col)
        elif CODE_HEADER_PATTERN.match(lowered) and col_num is None:
            col_num = col
            skip.add(col)
        elif LOCATION_HEADER_PATTERN.search(header):
            col_rack = col
            skip.add(col)
        elif QUANTITY_HEADER_PATTERN.search(header):
            col_qty = col
            skip.add(col)
        elif UNIT_HEADER_PATTERN.match(header):
            skip.add(col)
        elif NOTE_HEADER_PATTERN.search(header):
            skip.add(col)
        elif TYPE_HEADER_PATTERN.match(lowered):
            col_type = col
            skip.add(col)
        elif CODE_HEADER_PATTERN.match(lowered):
            if col != col_num:
                skip.add(col)
        elif any(token in lowered for token in TABLE_NAME_TOKENS):
            if col_name is None:
                col_name = col

    if col_num is None:
        col_num = 1
    if col_name is None:
        for col in range(1, max_col + 1):
            if col not in skip and col != col_num:
                col_name = col
                break
        if col_name is None:
            col_name = min(col_num + 1, max_col)

    candidates = set(range(DESCRIPTION_START_COL, DESCRIPTION_END_COL + 1))
    candidates.add(col_name)
    desc_cols: list[int] = []
    for col in sorted(candidates):
        if col == col_num or col in skip:
            continue
        if col_type is not None and col == col_type:
            continue
        desc_cols.append(col)
    if col_name not in desc_cols:
        desc_cols.insert(0, col_name)

    return TableLayout(
        col_num=col_num,
        col_name=col_name,
        col_type=col_type,
        col_qty=col_qty,
        col_rack=col_rack,
        desc_cols=tuple(desc_cols),
        skip_desc=frozenset(skip),
    )


def _detect_non_description_columns(worksheet: _SheetView, header_row: int) -> frozenset[int]:
    return _detect_table_layout(worksheet, header_row).skip_desc


def _is_main_row(
    worksheet: _SheetView,
    row: int,
    layout: TableLayout,
    *,
    require_fill: bool = True,
    line_number: int | None = None,
) -> bool:
    if _is_table_header_row(worksheet, row):
        return False
    if line_number is None:
        line_number = _integer_anchor(worksheet, row, layout)
    if line_number is None or line_number < 1:
        return False
    if not _merge_description(worksheet, row, layout):
        return False
    fill_col = layout.col_name
    if require_fill:
        return _is_main_fill(worksheet.cell_fill_rgb(row, fill_col))
    return True


def _integer_anchor(
    worksheet: _SheetView,
    row: int,
    layout: TableLayout,
) -> int | None:
    parsed = _parse_positive_int(worksheet.cell_value(row, layout.col_num))
    if parsed is not None:
        return parsed
    max_col = min(max(worksheet.max_column, 8), 12)
    for col in range(1, max_col + 1):
        text = _cell_text(worksheet, row, col)
        if not text:
            continue
        return _parse_positive_int(worksheet.cell_value(row, col))
    return None


def _parse_positive_int(value: object) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value >= 1 else None
    if isinstance(value, float):
        if not value.is_integer() or value < 1:
            return None
        return int(value)
    text = normalize_incoming_text(value)
    if not text.isdigit():
        return None
    parsed = int(text)
    return parsed if parsed >= 1 else None


def _read_row_quantity(worksheet: _SheetView, row: int, layout: TableLayout) -> int:
    quantity = _try_read_row_quantity(worksheet, row, layout)
    if quantity is None:
        raise ValueError(f"Quantity is missing in main row {row}")
    return quantity


def _try_read_row_quantity(worksheet: _SheetView, row: int, layout: TableLayout) -> Optional[int]:
    preferred: list[int] = []
    if layout.col_qty is not None:
        preferred.append(layout.col_qty)
    for col in QUANTITY_COLS:
        if col not in preferred:
            preferred.append(col)
    for col in preferred:
        value = worksheet.cell_value(row, col)
        if value is None or str(value).strip() == "":
            continue
        try:
            return _parse_quantity(value)
        except ValueError:
            continue
    return None


def _row_is_blank(worksheet: _SheetView, row: int) -> bool:
    max_col = min(worksheet.max_column, 12)
    return all(not _cell_text(worksheet, row, col) for col in range(1, max_col + 1))


def _is_alias_row(worksheet: _SheetView, row: int, layout: TableLayout) -> bool:
    for col in _alias_scan_cols(layout):
        if _is_alias_fill(worksheet.cell_fill_rgb(row, col)):
            return True
    alias_text = _row_name_text(worksheet, row, layout)
    return bool(alias_text) and alias_text.upper().startswith("IMP")


def _alias_scan_cols(layout: TableLayout) -> tuple[int, ...]:
    cols = [layout.col_name, 2, *layout.desc_cols]
    seen: set[int] = set()
    ordered: list[int] = []
    for col in cols:
        if col in seen:
            continue
        seen.add(col)
        ordered.append(col)
    return tuple(ordered)


def _is_main_fill(rgb: str) -> bool:
    if _fill_hex(rgb) in {MAIN_ROW_FILL, "E0FFE0"}:
        return True
    parsed = _rgb_tuple(rgb)
    if parsed is None:
        return False
    red, green, blue = parsed
    return green >= 200 and red <= 240 and blue <= 240 and green >= red and green >= blue


def _is_alias_fill(rgb: str) -> bool:
    if _fill_hex(rgb) in {ALIAS_ROW_FILL, "FFFFC0"}:
        return True
    parsed = _rgb_tuple(rgb)
    if parsed is None:
        return False
    red, green, blue = parsed
    return red >= 220 and green >= 200 and blue <= 220


def _fill_hex(rgb: str) -> str:
    return str(rgb).upper().replace("#", "")


def _rgb_tuple(rgb: str) -> tuple[int, int, int] | None:
    hex_color = _fill_hex(rgb)
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    if len(hex_color) != 6:
        return None
    try:
        return int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    except ValueError:
        return None


def _extract_service_line(worksheet: _SheetView, row: int, layout: TableLayout) -> Optional[str]:
    cols = list(dict.fromkeys((layout.col_name, *layout.desc_cols, *SERVICE_TEXT_COLS)))
    for col in cols:
        text = _cell_text(worksheet, row, col)
        if text and _looks_like_service_line(text):
            return text
    return None


def _looks_like_service_line(text: str) -> bool:
    lowered = text.lower()
    return (
        "продаж" in lowered
        or "заказ:" in lowered
        or "урл_" in lowered
        or "урп_" in lowered
        or "перемещение на склад" in lowered
    )


def _is_cell_location_token(text: str) -> bool:
    return bool(CELL_LOCATION_TOKEN_PATTERN.fullmatch(text.strip()))


def _row_name_text(worksheet: _SheetView, row: int, layout: TableLayout) -> str:
    for col in _alias_scan_cols(layout):
        text = _cell_text(worksheet, row, col)
        if not text or _is_cell_location_token(text):
            continue
        cleaned = sanitize_warehouse_topology(text)
        if cleaned:
            return cleaned
    return _merge_description(worksheet, row, layout)


def _merge_description(
    worksheet: _SheetView,
    row: int,
    layout: TableLayout,
) -> str:
    parts: list[str] = []
    for col in layout.desc_cols:
        if col in layout.skip_desc:
            continue
        text = _cell_text(worksheet, row, col)
        if not text:
            continue
        if _is_cell_location_token(text):
            continue
        if parts and re.fullmatch(r"\d+(?:[.,]\d+)?", text):
            continue
        cleaned = sanitize_warehouse_topology(text)
        if cleaned:
            parts.append(cleaned)
    return " ".join(parts)


def _cell_text(worksheet: _SheetView, row: int, col: int) -> str:
    value = worksheet.cell_value(row, col)
    if value is None:
        return ""
    return normalize_incoming_text(value)


def _parse_quantity(value: object) -> int:
    if value is None:
        raise ValueError("Quantity is missing in main row")
    if isinstance(value, bool):
        raise ValueError("Quantity must not be boolean")
    if isinstance(value, int):
        quantity = value
    elif isinstance(value, float):
        quantity = int(value)
    else:
        normalized = normalize_incoming_text(value).replace(",", ".")
        if not re.fullmatch(r"\d+(?:\.\d+)?", normalized):
            raise ValueError(f"Invalid quantity value: {value!r}")
        quantity = int(float(normalized))
    if quantity < 1:
        raise ValueError(f"Quantity must be >= 1, got {quantity}")
    return quantity
