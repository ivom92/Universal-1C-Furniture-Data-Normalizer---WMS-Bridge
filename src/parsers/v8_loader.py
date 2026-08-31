"""Loader for the 1C v8 master catalog Excel export (17 columns, 10k+ rows)."""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.models import CatalogEntity
from src.preprocessor.normalizer import canonicalize_dimensions

V8_COLUMN_HEADERS: tuple[str, ...] = (
    "Номенклатура",
    "ХарактеристикаНоменклатуры",
    "НоменклатураКод",
    "Штрихкод",
    "Вес",
    "Объем",
    "Высота",
    "Длина",
    "Глубина",
    "ЭтикеткаМодель",
    "Модуль",
    "Цвет",
    "Начинка",
    "Упаковка",
    "ТипЭтикетки",
    "ДС",
    "ЗонаХранения",
)

_SCIENTIFIC_RE = re.compile(r"^[+-]?\d[\d,]*(?:[.,]\d+)?[eE][+-]?\d+$")


def format_nomenclature_code(value: Any, number_format: str = "") -> str:
    """Preserve leading zeros for factory PK (e.g. 64794 -> '00000064794')."""
    if value is None:
        return ""

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return ""
        if text.isdigit() and number_format:
            width = _zero_pad_width(number_format)
            if width:
                return text.zfill(width)
        return text

    if isinstance(value, bool):
        raise ValueError("НоменклатураКод must not be boolean")

    if isinstance(value, int):
        numeric = value
    elif isinstance(value, float):
        numeric = int(value)
    else:
        return str(value).strip()

    width = _zero_pad_width(number_format)
    if width:
        return str(numeric).zfill(width)
    return str(numeric)


def restore_barcode(value: Any, number_format: str = "") -> Optional[str]:
    """
    Normalize EAN-13 from Excel cell values.

    Handles integer cells, spaced strings, and exponential corruption
    (e.g. ``2,006E+12`` -> ``2006000045445`` when full precision is available).
    """
    if value is None:
        return None

    if isinstance(value, bool):
        raise ValueError("Штрихкод must not be boolean")

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        return _float_to_integer_string(value)

    text = str(value).strip()
    if not text:
        return None

    compact = text.replace(" ", "")
    if _SCIENTIFIC_RE.match(compact):
        normalized = compact.replace(",", ".")
        try:
            decimal_value = Decimal(normalized)
        except InvalidOperation:
            return text
        return _decimal_to_integer_string(decimal_value)

    return text


def load_catalog_v8(path: str | Path) -> list[CatalogEntity]:
    """Load the full v8 catalog workbook into validated ``CatalogEntity`` rows."""
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Catalog file not found: {workbook_path}")

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        worksheet = workbook.active
        _validate_header_row(worksheet)
        return [_row_to_entity(row_index, row_cells) for row_index, row_cells in _iter_data_rows(worksheet)]
    finally:
        workbook.close()


def _validate_header_row(worksheet: Worksheet) -> None:
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, max_col=17, values_only=False))
    headers = [cell.value for cell in header_cells]
    if headers != list(V8_COLUMN_HEADERS):
        raise ValueError(
            "Unexpected v8 catalog header. "
            f"Expected {V8_COLUMN_HEADERS}, got {tuple(headers)}"
        )


def _iter_data_rows(worksheet: Worksheet):
    for row_index, row_cells in enumerate(
        worksheet.iter_rows(min_row=2, max_col=17, values_only=False),
        start=2,
    ):
        if _row_is_empty(row_cells):
            continue
        yield row_index, row_cells


def _row_is_empty(cells: tuple[Any, ...]) -> bool:
    return all(cell.value is None or str(cell.value).strip() == "" for cell in cells)


def _row_to_entity(row_index: int, cells: tuple[Any, ...]) -> CatalogEntity:
    values = {
        "Номенклатура": _as_optional_str(cells[0].value) or "",
        "ХарактеристикаНоменклатуры": _as_optional_str(cells[1].value),
        "НоменклатураКод": format_nomenclature_code(cells[2].value, cells[2].number_format),
        "Штрихкод": restore_barcode(cells[3].value, cells[3].number_format),
        "Вес": _as_optional_float(cells[4].value),
        "Объем": _as_optional_float(cells[5].value),
        "Высота": _as_optional_float(cells[6].value),
        "Длина": _as_optional_float(cells[7].value),
        "Глубина": _as_optional_float(cells[8].value),
        "ЭтикеткаМодель": _as_optional_str(cells[9].value),
        "Модуль": _as_optional_str(cells[10].value),
        "Цвет": _as_optional_str(cells[11].value),
        "Начинка": _as_optional_str(cells[12].value),
        "Упаковка": _as_optional_str(cells[13].value),
        "ТипЭтикетки": _as_optional_str(cells[14].value),
        "ДС": _as_optional_str(cells[15].value),
        "ЗонаХранения": _as_optional_str(cells[16].value),
    }

    if not values["Номенклатура"]:
        raise ValueError(f"Row {row_index}: empty Номенклатура")
    if not values["НоменклатураКод"]:
        raise ValueError(f"Row {row_index}: empty НоменклатураКод")

    return CatalogEntity.model_validate(values)


def _zero_pad_width(number_format: str) -> int:
    if not number_format:
        return 0
    if set(number_format) <= {"0"}:
        return len(number_format)
    return 0


def _float_to_integer_string(value: float) -> str:
    decimal_value = Decimal(str(value))
    return _decimal_to_integer_string(decimal_value)


def _decimal_to_integer_string(value: Decimal) -> str:
    integral = value.to_integral_value()
    return format(integral, "f")


def _as_optional_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


_DIMENSION_RUN_RE = re.compile(r"\d+(?:\s*[хxX×]\s*\d+)+")
_DOTTED_ARTICLE_RE = re.compile(r"\b\d+(?:\.\d+){2,}\b")
_STAR_ARTICLE_RE = re.compile(r"\b\d{3,}\*\d{2,}\b")
_PLAIN_ARTICLE_RE = re.compile(r"\d{4,}")
_HARDWARE_TYPE_RE = re.compile(
    r"корнер|планка|плинтус|профиль|заглушка|светильник|"
    r"\bопора\b|\bпетл(?:я|и|ю)\b|\bручк(?:а|и|у)\b|стяжк[а-яё]*|"
    r"\bнавес\b|\bугол\b|доводчик|направляющ|полкодерж",
    re.IGNORECASE,
)


def extract_article_tokens(text: str) -> list[str]:
    """Extract catalog article / technical-number tokens (length ≥ 4, dotted SKUs, star sizes)."""
    if not text:
        return []

    text = canonicalize_dimensions(text)
    tokens: list[str] = []
    seen: set[str] = set()

    def _add(raw: str) -> None:
        token = raw.strip().lower().replace(" ", "")
        if token and token not in seen:
            seen.add(token)
            tokens.append(token)

    for match in _DOTTED_ARTICLE_RE.finditer(text):
        _add(match.group(0))
    for match in _STAR_ARTICLE_RE.finditer(text):
        _add(match.group(0))

    masked = _DIMENSION_RUN_RE.sub(" ", text)
    masked = _DOTTED_ARTICLE_RE.sub(" ", masked)
    masked = _STAR_ARTICLE_RE.sub(" ", masked)
    for match in _PLAIN_ARTICLE_RE.finditer(masked):
        _add(match.group(0))

    return tokens


def extract_hardware_types(text: str) -> list[str]:
    """Return hardware type tokens present in text (from catalog-driven matching, not furniture models)."""
    if not text:
        return []
    lowered = text.lower()
    seen: set[str] = set()
    types: list[str] = []
    for match in _HARDWARE_TYPE_RE.finditer(lowered):
        token = match.group(0).lower()
        if token == "угол" and "цокол" not in lowered and not re.search(r"\d+\s*гр", lowered):
            continue
        if token not in seen:
            seen.add(token)
            types.append(token)
    return types


def build_article_index(catalog: list[CatalogEntity]) -> dict[str, list[CatalogEntity]]:
    """Map exact article tokens found in v8 nomenclature to catalog rows."""
    index: dict[str, list[CatalogEntity]] = {}
    seen_codes: dict[str, set[str]] = {}
    for entity in catalog:
        blob = " ".join(
            part
            for part in (
                entity.nomenclature,
                entity.module,
                entity.filling,
                entity.label_type,
                entity.characteristic,
            )
            if part
        )
        for token in extract_article_tokens(blob):
            used = seen_codes.setdefault(token, set())
            if entity.nomenclature_code in used:
                continue
            used.add(entity.nomenclature_code)
            index.setdefault(token, []).append(entity)
    return index


def build_hardware_type_index(catalog: list[CatalogEntity]) -> dict[str, list[CatalogEntity]]:
    """Map hardware type words present in the loaded catalog to matching rows."""
    index: dict[str, list[CatalogEntity]] = {}
    seen_codes: dict[str, set[str]] = {}
    for entity in catalog:
        blob = entity.nomenclature or ""
        for token in extract_hardware_types(blob):
            used = seen_codes.setdefault(token, set())
            if entity.nomenclature_code in used:
                continue
            used.add(entity.nomenclature_code)
            index.setdefault(token, []).append(entity)
    return index


def _as_optional_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return None
