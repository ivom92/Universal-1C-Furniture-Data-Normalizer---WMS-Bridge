"""Builders for synthetic 1C v7.7 picking-list workbooks (.xlsx and .xls)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

MAIN_FILL = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
ALIAS_FILL = PatternFill(start_color="FFFFC0", end_color="FFFFC0", fill_type="solid")

SAMPLE_CUSTOMER = "ИП Тестов Алексей"
SAMPLE_BLOCKS = (
    {
        "line_number": 1,
        "description": "КДР к Столешница 3000х600х40 Дуб сонома",
        "item_type": "",
        "quantity": 1,
        "alias": "КДР к Столешница 3000х600х40",
        "service": "Продажи оптовые УРП_001 Заказ: ЦНТ-100",
    },
    {
        "line_number": 2,
        "description": "Стекло к Витрина 116х596",
        "item_type": "Стекло",
        "quantity": 2,
        "alias": "IMP ст Витрина 116х596",
        "service": "Продажи оптовые УРП_002 Заказ: ЦНТ-100",
    },
)


def write_sample_v7_xlsx(
    path: Path,
    customer_label: str = "Покупатель:",
    customer_name: str = SAMPLE_CUSTOMER,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отборочная"
    sheet.cell(1, 1, "Отборочный лист")
    sheet.cell(2, 1, customer_label)
    sheet.cell(2, 2, customer_name)

    row = 8
    for block in SAMPLE_BLOCKS:
        _write_xlsx_block(sheet, row, block)
        row += 3

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def write_sample_v7_xls(
    path: Path,
    customer_label: str = "Покупатель:",
    customer_name: str = SAMPLE_CUSTOMER,
) -> Path:
    import xlwt
    from xlwt.Style import add_palette_colour

    add_palette_colour("v7_green", 0x21)
    add_palette_colour("v7_yellow", 0x22)

    workbook = xlwt.Workbook()
    workbook.set_colour_RGB(0x21, 0xE0, 0xFF, 0xE0)
    workbook.set_colour_RGB(0x22, 0xFF, 0xFF, 0xC0)
    sheet = workbook.add_sheet("Отборочная")

    green = xlwt.easyxf("pattern: pattern solid, fore_colour v7_green")
    yellow = xlwt.easyxf("pattern: pattern solid, fore_colour v7_yellow")

    sheet.write(0, 0, "Отборочный лист")
    sheet.write(1, 0, customer_label)
    sheet.write(1, 1, customer_name)

    row = 7
    for block in SAMPLE_BLOCKS:
        _write_xls_block(sheet, row, block, green, yellow)
        row += 3

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
    return path


def _write_xlsx_block(sheet, start_row: int, block: dict) -> None:
    sheet.cell(start_row, 1, block["line_number"])
    desc_cell = sheet.cell(start_row, 2, block["description"])
    desc_cell.fill = MAIN_FILL
    sheet.cell(start_row, 7, block["item_type"])
    sheet.cell(start_row, 8, block["quantity"])

    alias_cell = sheet.cell(start_row + 1, 2, block["alias"])
    alias_cell.fill = ALIAS_FILL
    sheet.cell(start_row + 2, 2, block["service"])


def _write_xls_block(sheet, start_row: int, block: dict, green, yellow) -> None:
    sheet.write(start_row, 0, block["line_number"])
    sheet.write(start_row, 1, block["description"], green)
    sheet.write(start_row, 6, block["item_type"])
    sheet.write(start_row, 7, block["quantity"])
    sheet.write(start_row + 1, 1, block["alias"], yellow)
    sheet.write(start_row + 2, 1, block["service"])


def write_transfer_v7_xlsx(
    path: Path,
    *,
    recipient: str = "Челябинск ТК",
    inline_header: bool = True,
    two_row_items: bool = False,
    include_recipient_label: bool = True,
    document_title: str = "Перемещение № ЧЛ-00452 от 01.09.2026",
) -> Path:
    """Synthetic 1C transfer (перемещение) print form."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Перемещение"
    sheet.cell(1, 1, document_title)

    if include_recipient_label:
        if inline_header:
            sheet.merge_cells("A2:D2")
            sheet.cell(2, 1, f"Склад-получатель: {recipient}")
        else:
            sheet.cell(2, 1, "Склад-получатель:")
            sheet.cell(2, 2, recipient)

    sheet.cell(5, 1, "№")
    sheet.cell(5, 2, "Код")
    sheet.cell(5, 3, "Товар")
    sheet.cell(5, 4, "Количество")
    sheet.cell(5, 5, "Ед.")

    row = 6
    for index, block in enumerate(SAMPLE_BLOCKS, start=1):
        sheet.cell(row, 1, index)
        desc_cell = sheet.cell(row, 2, block["description"])
        desc_cell.fill = MAIN_FILL
        sheet.cell(row, 3, block["description"])
        sheet.cell(row, 4, block["quantity"])
        sheet.cell(row, 7, block["item_type"])
        sheet.cell(row, 8, block["quantity"])
        if two_row_items:
            sheet.cell(row + 1, 2, block["service"])
            row += 2
        else:
            alias_cell = sheet.cell(row + 1, 2, block["alias"])
            alias_cell.fill = ALIAS_FILL
            sheet.cell(row + 2, 2, block["service"])
            row += 3

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def write_location_topology_v7_xlsx(
    path: Path,
    *,
    warehouse: str = "РС УрФО Империал",
    inline_warehouse: bool = True,
    first_item: str = "Аврора Зеркало 1/1 венге",
    cell_token: str = "Р1.16.Я2",
) -> Path:
    """Transfer-style sheet with warehouse cell topology in column B."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Перемещение"
    sheet.cell(1, 1, "Перемещение № ЧЛ-00452 от 01.09.2026")
    if inline_warehouse:
        sheet.cell(2, 1, f"Рег.склад: {warehouse}")
    else:
        sheet.cell(2, 1, "Рег.склад:")
        sheet.cell(2, 2, warehouse)

    sheet.cell(5, 1, "№")
    sheet.cell(5, 2, "Линейка / Секция / Место")
    sheet.cell(5, 3, "Упаковка + цвет")
    sheet.cell(5, 4, "Количество")
    sheet.cell(5, 5, "Ед.")

    sheet.cell(6, 1, 1)
    loc_cell = sheet.cell(6, 2, cell_token)
    loc_cell.fill = MAIN_FILL
    name_cell = sheet.cell(6, 3, first_item)
    name_cell.fill = MAIN_FILL
    sheet.cell(6, 4, 1)
    sheet.cell(6, 8, 1)
    sheet.cell(7, 2, "Продажи оптовые УРП_001 Заказ: ЦНТ-100")

    sheet.cell(8, 1, 2)
    loc_cell_2 = sheet.cell(8, 2, "Р10.12.Я1")
    loc_cell_2.fill = MAIN_FILL
    name_cell_2 = sheet.cell(8, 3, "Аврора Кровать 140 с основанием 1/2 дуб сонома/белый")
    name_cell_2.fill = MAIN_FILL
    sheet.cell(8, 4, 1)
    sheet.cell(8, 8, 1)
    sheet.cell(9, 2, "Продажи оптовые УРП_002 Заказ: ЦНТ-100")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def write_shifted_columns_v7_xlsx(path: Path) -> Path:
    """Printed form with columns shifted right and extra service rows between items."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отборочная"
    sheet.cell(1, 1, "Отборочный лист")
    sheet.cell(2, 1, "Покупатель:")
    sheet.cell(2, 2, SAMPLE_CUSTOMER)

    sheet.cell(5, 2, "№")
    sheet.cell(5, 3, "Секция")
    sheet.cell(5, 4, "Наименование")
    sheet.cell(5, 5, "Тип")
    sheet.cell(5, 6, "Кол-во")
    sheet.cell(5, 7, "Отметка")

    sheet.cell(6, 2, 1)
    loc = sheet.cell(6, 3, "Р1.1")
    loc.fill = MAIN_FILL
    name = sheet.cell(6, 4, SAMPLE_BLOCKS[0]["description"])
    name.fill = MAIN_FILL
    sheet.cell(6, 5, SAMPLE_BLOCKS[0]["item_type"])
    sheet.cell(6, 6, SAMPLE_BLOCKS[0]["quantity"])
    alias = sheet.cell(7, 4, SAMPLE_BLOCKS[0]["alias"])
    alias.fill = ALIAS_FILL
    sheet.cell(8, 4, "примечание склада")
    sheet.cell(9, 4, SAMPLE_BLOCKS[0]["service"])

    sheet.cell(11, 2, 2)
    loc2 = sheet.cell(11, 3, "Р2.2")
    loc2.fill = MAIN_FILL
    name2 = sheet.cell(11, 4, SAMPLE_BLOCKS[1]["description"])
    name2.fill = MAIN_FILL
    sheet.cell(11, 5, SAMPLE_BLOCKS[1]["item_type"])
    sheet.cell(11, 6, SAMPLE_BLOCKS[1]["quantity"])
    alias2 = sheet.cell(12, 4, SAMPLE_BLOCKS[1]["alias"])
    alias2.fill = ALIAS_FILL
    sheet.cell(13, 4, SAMPLE_BLOCKS[1]["service"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def write_five_row_block_v7_xlsx(path: Path) -> Path:
    """Item blocks of mixed height (1 / 5 rows) with glued warehouse tokens."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отборочная"
    sheet.cell(1, 1, "Отборочный лист")
    sheet.cell(2, 1, "Покупатель:")
    sheet.cell(2, 2, SAMPLE_CUSTOMER)

    sheet.cell(4, 1, "№ п/п")
    sheet.cell(4, 2, "Адрес")
    sheet.cell(4, 3, "Номенклатура")
    sheet.cell(4, 4, "К-во")

    sheet.cell(5, 1, 1)
    name = sheet.cell(5, 3, "Секция 12 Аврора Зеркало 1/1 венге")
    name.fill = MAIN_FILL
    sheet.cell(5, 4, 1)

    sheet.cell(7, 1, 2)
    loc = sheet.cell(7, 2, "Р3.4.Я1")
    loc.fill = MAIN_FILL
    name2 = sheet.cell(7, 3, SAMPLE_BLOCKS[0]["description"])
    name2.fill = MAIN_FILL
    sheet.cell(7, 4, SAMPLE_BLOCKS[0]["quantity"])
    alias = sheet.cell(8, 3, SAMPLE_BLOCKS[0]["alias"])
    alias.fill = ALIAS_FILL
    sheet.cell(9, 3, "примечание комплектовки")
    sheet.cell(10, 3, "Перемещение на склад УРФО")
    sheet.cell(11, 3, SAMPLE_BLOCKS[0]["service"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def write_soft_furniture_xlsx(path: Path) -> Path:
    """Soft-furniture transfer: one parent SKU expanded into N packages."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отборочная"
    sheet.cell(1, 1, "Отборочная ведомость. Мягкая мебель")
    sheet.cell(2, 1, "КМР_ПС07027 (Заказ: ЦНТ-001513)")
    sheet.cell(4, 1, "№")
    sheet.cell(4, 2, "Наименование")
    sheet.cell(4, 3, "Количество")

    sheet.cell(5, 1, 1)
    sheet.cell(5, 2, "SF сп Кровать Вена с под мех")
    sheet.cell(6, 2, "Ткань: shadow")
    sheet.cell(7, 2, "Состоит из упаковок:")
    sheet.cell(8, 2, "Изголовье/Дно бельевого короба 1/3")
    sheet.cell(8, 3, 1)
    sheet.cell(9, 2, "Каркас 2/3")
    sheet.cell(9, 3, 1)
    sheet.cell(10, 2, "Матрас 3/3")
    sheet.cell(10, 3, 1)
    sheet.cell(12, 1, "ИТОГО мест по отборочной ведомости")
    sheet.cell(12, 3, 3)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def write_soft_furniture_multi_xlsx(path: Path) -> Path:
    """Three parents: sofa 2 packs, bed 3 packs, pouf monoblock → 6 places."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отборочная"
    sheet.cell(1, 1, "Отборочная ведомость. Мягкая мебель")
    sheet.cell(2, 1, "Перемещение: КМР_А, КМР_Б, КМР_В")
    sheet.cell(4, 1, "№ п/п")
    sheet.cell(4, 2, "Товар")
    sheet.cell(4, 5, "Ткань")
    sheet.cell(4, 6, "Цвет")
    sheet.cell(4, 7, "Кол-во")
    sheet.cell(4, 8, "Вес, кг")
    sheet.cell(4, 9, "Отметка")

    sheet.cell(5, 1, 1)
    sheet.cell(5, 2, "Диван Милан")
    sheet.cell(5, 6, "velvet")
    sheet.cell(5, 7, 1)
    sheet.cell(6, 5, "Перемещение на склад КМР_А (Заказ: ЦНТ-100)")
    sheet.cell(7, 4, "Состоит из упаковок:")
    sheet.cell(7, 5, "Сиденье 1/2")
    sheet.cell(7, 7, 1)
    sheet.cell(8, 5, "Спинка 2/2")
    sheet.cell(8, 7, 1)

    sheet.cell(9, 1, 2)
    sheet.cell(9, 2, "Кровать Вена")
    sheet.cell(9, 6, "loft")
    sheet.cell(9, 7, 1)
    sheet.cell(10, 5, "Перемещение на склад КМР_Б (Заказ: ЦНТ-200)")
    sheet.cell(11, 4, "Состоит из упаковок:")
    sheet.cell(11, 5, "Изголовье 1/3")
    sheet.cell(11, 7, 1)
    sheet.cell(12, 5, "Боковины 2/3")
    sheet.cell(12, 7, 1)
    sheet.cell(13, 5, "Основание 3/3")
    sheet.cell(13, 7, 1)

    sheet.cell(14, 1, 3)
    sheet.cell(14, 2, "Пуф Куб")
    sheet.cell(14, 6, "SF")
    sheet.cell(14, 7, 1)
    sheet.cell(15, 5, "Перемещение на склад КМР_В (Заказ: ЦНТ-300)")

    sheet.cell(17, 6, "ИТОГО мест по отборочной ведомости")
    sheet.cell(17, 7, 6)
    sheet.cell(19, 2, "Кладовщик __________________")
    sheet.cell(19, 6, "Экспедитор ______________________")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path

