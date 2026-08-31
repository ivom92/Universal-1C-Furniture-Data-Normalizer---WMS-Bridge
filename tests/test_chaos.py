"""Chaos / zero-loss tests for HTML-XLS, NFKC, and synthetic 1C layouts."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.adapters.wms_excel_adapter import WMS_COLUMNS, WMSExcelAdapter
from src.matcher.hybrid_matcher import HybridMatcher
from src.parsers.v7_parser import parse_v7_order
from tests.conftest import CATALOG_V8_PATH

pytest.importorskip("lxml")


def _assert_monotonic_line_numbers(path: Path, hybrid_matcher: HybridMatcher) -> int:
    parsed = parse_v7_order(path)
    assert parsed.blocks, f"Parser returned no blocks for {path.name}"
    source_numbers = [block.order_line_number for block in parsed.blocks]
    assert source_numbers == list(range(1, len(parsed.blocks) + 1)), source_numbers

    decisions = hybrid_matcher.match_order_decisions(parsed.blocks)
    assert len(decisions) == len(parsed.blocks)
    exported = WMSExcelAdapter().export_to_bytes(decisions, parsed.customer_name)
    workbook = load_workbook(exported)
    worksheet = workbook.active
    data_last_row = WMSExcelAdapter.wms_data_last_row(len(parsed.blocks))
    data_rows = data_last_row - 1
    headers = [worksheet.cell(row=1, column=col).value for col in range(1, 6)]
    excel_numbers = [
        worksheet.cell(row=row, column=1).value for row in range(2, data_last_row + 1)
    ]
    workbook.close()
    assert headers == WMS_COLUMNS
    assert data_rows == len(parsed.blocks)
    assert excel_numbers == list(range(1, data_rows + 1))
    assert all(isinstance(value, int) and not isinstance(value, bool) for value in excel_numbers)
    return data_rows


def _run_pipeline(path: Path, hybrid_matcher: HybridMatcher) -> tuple[int, int]:
    parsed = parse_v7_order(path)
    assert parsed.blocks, f"Parser returned no blocks for {path.name}"
    rows = _assert_monotonic_line_numbers(path, hybrid_matcher)
    return len(parsed.blocks), rows


@pytest.mark.skipif(not CATALOG_V8_PATH.exists(), reason="Real catalog file is missing in data/")
class TestSyntheticChaosOrders:
    def test_named_generator_files_are_zero_loss(
        self,
        tmp_path: Path,
        catalog_v8,
        hybrid_matcher: HybridMatcher,
    ) -> None:
        from scripts.generate_synthetic_orders import generate_named_fixtures

        paths = generate_named_fixtures(tmp_path, catalog_v8)
        assert len(paths) == 4
        for path in paths:
            blocks, rows = _run_pipeline(path, hybrid_matcher)
            assert blocks == rows
            assert blocks >= 1

    def test_html_skipped_cells_keep_monotonic_line_numbers(
        self,
        tmp_path: Path,
        catalog_v8,
        hybrid_matcher: HybridMatcher,
    ) -> None:
        from scripts.generate_synthetic_orders import (
            sample_catalog_items,
            write_html_xls_with_skipped_cells,
        )

        items = sample_catalog_items(catalog_v8, count=8)
        path = write_html_xls_with_skipped_cells(tmp_path / "skipped_cells.xls", items)
        _assert_monotonic_line_numbers(path, hybrid_matcher)

    def test_twenty_layout_mutations_never_crash(
        self,
        tmp_path: Path,
        catalog_v8,
        hybrid_matcher: HybridMatcher,
    ) -> None:
        from scripts.generate_synthetic_orders import sample_catalog_items, write_chaos_variant

        items = sample_catalog_items(catalog_v8, count=30)
        variants: list[Path] = []
        for index in range(20):
            path = tmp_path / f"chaos_{index}"
            variants.append(
                write_chaos_variant(
                    path,
                    items,
                    shift_rows=3 + (index % 8),
                    as_html=index % 3 == 0,
                    skip_aliases=index % 3 == 1,
                    star_dimensions=index % 3 == 2,
                )
            )
        assert len(variants) >= 20
        for path in variants:
            _run_pipeline(path, hybrid_matcher)
