"""End-to-end integration test: v7.7 order → hybrid match → WMS Excel export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.adapters.wms_excel_adapter import WMS_COLUMNS, WMSExcelAdapter
from src.matcher.hybrid_matcher import HybridMatcher

EXPECTED_CUSTOMER = "Рубан Кристина Олеговна ИП"
EXPECTED_DATA_ROWS = 55


class TestEndToEndPipeline:
    def test_full_pipeline_v7_to_wms_excel(
        self,
        hybrid_matcher: HybridMatcher,
        order_ruban_parsed,
        tmp_path: Path,
    ) -> None:
        blocks = order_ruban_parsed.blocks
        customer_name = order_ruban_parsed.customer_name

        assert customer_name == EXPECTED_CUSTOMER
        assert len(blocks) == EXPECTED_DATA_ROWS

        decisions = hybrid_matcher.match_order_decisions(blocks)
        assert len(decisions) == len(blocks)

        output_path = tmp_path / "wms_export.xlsx"
        adapter = WMSExcelAdapter()
        saved_path = adapter.export(decisions, customer_name, output_path)

        assert saved_path.exists()

        workbook = load_workbook(saved_path, data_only=False)
        worksheet = workbook.active

        data_last_row = WMSExcelAdapter.wms_data_last_row(EXPECTED_DATA_ROWS)
        assert worksheet.max_row == WMSExcelAdapter.wms_totals_row(EXPECTED_DATA_ROWS)
        assert worksheet.title == "Импорт_WMS"

        headers = [worksheet.cell(row=1, column=col).value for col in range(1, 6)]
        assert headers == WMS_COLUMNS

        matched_barcodes_checked = 0

        for row_index in range(2, data_last_row + 1):
            customer_cell = worksheet.cell(row=row_index, column=5)
            assert customer_cell.value == EXPECTED_CUSTOMER

            number_cell = worksheet.cell(row=row_index, column=1)
            nomenclature_cell = worksheet.cell(row=row_index, column=2)
            barcode_cell = worksheet.cell(row=row_index, column=3)
            quantity_cell = worksheet.cell(row=row_index, column=4)

            decision = decisions[row_index - 2]
            block = decision.raw_block

            assert int(number_cell.value) == block.order_line_number
            assert int(number_cell.value) == row_index - 1
            assert int(quantity_cell.value) == block.quantity

            if decision.status in {"MATCHED_AUTO", "MATCHED_LLM"}:
                assert decision.matched_entity is not None
                assert nomenclature_cell.value == decision.matched_entity.nomenclature
                if decision.matched_entity.barcode:
                    barcode_text = str(barcode_cell.value)
                    assert barcode_cell.data_type == "s"
                    assert barcode_text.isdigit()
                    assert len(barcode_text) == 13
                    matched_barcodes_checked += 1
                else:
                    assert barcode_cell.value in (None, "")
            else:
                assert barcode_cell.value in (None, "")

        assert matched_barcodes_checked >= 40
