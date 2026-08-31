"""Unit tests for shift history persistence and F5 session restore."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from src.adapters.wms_excel_adapter import WMS_COLUMNS, WMS_SHEET_NAME, WMSExcelAdapter
from src.models import CatalogEntity, ExtractedFeatures, MatchDecision, RawOrderBlock
from src.utils.history_manager import (
    DOC_TYPE_CABINET,
    DOC_TYPE_SOFT,
    HistoryManager,
    OrderRunMeta,
    build_order_run_meta,
    dump_decisions_for_session,
    load_decisions_from_session,
    shift_summary,
)


def _clock(stamp: str) -> datetime:
    return datetime.strptime(stamp, "%Y-%m-%d %H:%M:%S")


def _block(*, description: str, quantity: int = 2, line_number: int = 1) -> RawOrderBlock:
    return RawOrderBlock(
        line_number=line_number,
        client_description=description,
        item_type="Пачка",
        quantity=quantity,
        factory_alias=description,
        order_service_line="Продажи оптовые УРП_ test",
        excel_row_start=line_number,
    )


def _entity(*, nomenclature: str, barcode: str | None) -> CatalogEntity:
    return CatalogEntity.model_validate(
        {
            "Номенклатура": nomenclature,
            "НоменклатураКод": "00000010001",
            "Штрихкод": barcode,
            "Упаковка": "1/1",
        }
    )


def _decision(
    *,
    description: str,
    quantity: int,
    status: str,
    entity: CatalogEntity | None,
    line_number: int = 1,
    match_method: str | None = None,
) -> MatchDecision:
    if match_method is None:
        match_method = "vector_auto" if entity is not None else "AUTO_NO_BARCODE"
    return MatchDecision(
        raw_block=_block(description=description, quantity=quantity, line_number=line_number),
        extracted_features=ExtractedFeatures(),
        status=status,
        matched_entity=entity,
        confidence_score=1.0,
        match_method=match_method,
    )


def _cabinet_decisions() -> list[MatchDecision]:
    return [
        _decision(
            description="Кухня",
            quantity=2,
            status="MATCHED_AUTO",
            entity=_entity(nomenclature="Фабрика Кухня 1/1", barcode="2006000045445"),
            line_number=1,
        ),
        _decision(
            description="Полка",
            quantity=3,
            status="MATCHED_AUTO",
            entity=_entity(nomenclature="Полка стекло", barcode=None),
            line_number=2,
            match_method="AUTO_NO_BARCODE",
        ),
        _decision(
            description="Зеркало заказное",
            quantity=1,
            status="QUARANTINE",
            entity=None,
            line_number=3,
        ),
    ]


def _soft_decisions() -> list[MatchDecision]:
    return [
        _decision(
            description="Кровать Palermo 160",
            quantity=1,
            status="MATCHED_AUTO",
            entity=None,
            line_number=1,
            match_method="AUTO_NO_BARCODE",
        ),
        _decision(
            description="Кровать Palermo 160 2/2",
            quantity=1,
            status="MATCHED_AUTO",
            entity=None,
            line_number=2,
            match_method="AUTO_NO_BARCODE",
        ),
    ]


def _meta(**overrides: object) -> OrderRunMeta:
    payload = {
        "order_id": "abc123def4567890",
        "timestamp": "2026-08-30 10:15:00",
        "original_filename": "order_cabinet.xls",
        "doc_type": DOC_TYPE_CABINET,
        "total_rows": 10,
        "total_places": 20,
        "matched_auto_count": 8,
        "auto_no_barcode_count": 1,
        "quarantine_count": 1,
        "wms_excel_path": "",
        "customer_name": "ИП Тестов",
    }
    payload.update(overrides)
    return OrderRunMeta.model_validate(payload)


def test_save_run_writes_excel_and_manifest(tmp_path: Path) -> None:
    manager = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 10:15:00"))
    excel = b"PK\x03\x04fake-xlsx-bytes"
    saved = manager.save_run(_meta(), excel)

    assert saved.is_file()
    assert saved.name.startswith("WMS_20260830_101500_")
    assert saved.read_bytes() == excel

    history = manager.get_shift_history()
    assert len(history) == 1
    assert history[0].original_filename == "order_cabinet.xls"
    assert history[0].wms_excel_path.endswith(saved.name)
    assert (tmp_path / "2026-08-30" / "shift_manifest.json").is_file()


def test_history_newest_first_and_get_last_run(tmp_path: Path) -> None:
    manager = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 12:00:00"))
    manager.save_run(_meta(order_id="first", timestamp="2026-08-30 09:00:00"), b"one")
    manager.save_run(
        _meta(
            order_id="second",
            timestamp="2026-08-30 11:30:00",
            original_filename="order_bed.xls",
            doc_type=DOC_TYPE_SOFT,
        ),
        b"two",
    )

    history = manager.get_shift_history()
    assert [item.order_id for item in history] == ["second", "first"]
    last = manager.get_last_run()
    assert last is not None
    assert last.order_id == "second"
    assert last.original_filename == "order_bed.xls"


def test_create_shift_zip_contains_all_excel(tmp_path: Path) -> None:
    manager = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 12:00:00"))
    manager.save_run(_meta(order_id="a", timestamp="2026-08-30 09:00:00"), b"excel-a")
    manager.save_run(
        _meta(order_id="b", timestamp="2026-08-30 10:00:00", original_filename="bed.xls"),
        b"excel-b",
    )

    raw = manager.create_shift_zip()
    with ZipFile(BytesIO(raw)) as archive:
        names = archive.namelist()
        assert len(names) == 2
        assert all(name.endswith(".xlsx") for name in names)
        payloads = {archive.read(name) for name in names}
        assert payloads == {b"excel-a", b"excel-b"}


def test_f5_restore_reads_last_excel_without_pipeline(tmp_path: Path) -> None:
    """New HistoryManager instance simulates a Streamlit F5 / new session."""
    first = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 10:15:00"))
    payload = b"PK\x03\x04persisted-wms"
    first.save_run(_meta(), payload)

    restored_session = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 10:16:00"))
    last = restored_session.get_last_run()
    assert last is not None
    assert restored_session.read_excel_bytes(last) == payload
    assert last.total_rows == 10
    assert last.customer_name == "ИП Тестов"


def test_session_sidecar_roundtrip_decisions(tmp_path: Path) -> None:
    manager = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 10:15:00"))
    decisions = _cabinet_decisions()
    session = {
        "decisions": dump_decisions_for_session(decisions),
        "customer_name": "ИП Тестов",
        "operator_overrides": {2: "1234567890123"},
        "upload_name": "order_cabinet.xls",
    }
    saved = manager.save_run(_meta(order_id="sess1"), b"xlsx", session=session)
    last = manager.get_last_run()
    assert last is not None
    loaded = manager.load_session(last)
    assert loaded is not None
    restored = load_decisions_from_session(loaded["decisions"])
    assert len(restored) == 3
    assert restored[0].raw_block.client_description == "Кухня"
    assert loaded["operator_overrides"][2] == "1234567890123"
    sidecar = saved.parent / f"{saved.stem}.session.json"
    assert sidecar.is_file()


def test_batch_two_independent_wms_files(tmp_path: Path) -> None:
    adapter = WMSExcelAdapter()
    manager = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 18:00:00"))

    cabinet = _cabinet_decisions()
    cabinet_bytes = adapter.export_to_bytes(
        cabinet,
        "РС УрФО Империал",
        source_name="order_transfering_01_09.xls",
    ).getvalue()
    cabinet_meta = build_order_run_meta(
        original_filename="order_transfering_01_09.xls",
        doc_type=DOC_TYPE_CABINET,
        decisions=cabinet,
        excel_bytes=cabinet_bytes,
        customer_name="РС УрФО Империал",
        timestamp="2026-08-30 18:00:01",
        order_id="batch-cabinet",
    )
    manager.save_run(cabinet_meta, cabinet_bytes)

    soft = _soft_decisions()
    soft_bytes = adapter.export_to_bytes(
        soft,
        "Склад мягкой мебели",
        source_name="order_transfering_01_09_bed.xls",
    ).getvalue()
    soft_meta = build_order_run_meta(
        original_filename="order_transfering_01_09_bed.xls",
        doc_type=DOC_TYPE_SOFT,
        decisions=soft,
        excel_bytes=soft_bytes,
        customer_name="Склад мягкой мебели",
        timestamp="2026-08-30 18:00:02",
        order_id="batch-soft",
    )
    manager.save_run(soft_meta, soft_bytes)

    history = manager.get_shift_history()
    assert len(history) == 2
    assert {item.original_filename for item in history} == {
        "order_transfering_01_09.xls",
        "order_transfering_01_09_bed.xls",
    }
    assert {item.doc_type for item in history} == {DOC_TYPE_CABINET, DOC_TYPE_SOFT}

    for item in history:
        workbook = load_workbook(BytesIO(manager.read_excel_bytes(item)), data_only=True)
        assert WMS_SHEET_NAME in workbook.sheetnames
        sheet = workbook[WMS_SHEET_NAME]
        headers = [sheet.cell(1, col).value for col in range(1, 6)]
        assert headers == WMS_COLUMNS
        data_rows = item.total_rows
        assert sheet.cell(2, 1).value is not None
        assert sheet.cell(data_rows + 1, 1).value is not None

    zip_raw = manager.create_shift_zip()
    with ZipFile(BytesIO(zip_raw)) as archive:
        assert len(archive.namelist()) == 2

    summary = shift_summary(history)
    assert summary["orders"] == 2
    assert summary["places"] == 6 + 2
    assert summary["quarantine"] == 1


def test_scan_override_updates_only_selected_order_excel(tmp_path: Path) -> None:
    adapter = WMSExcelAdapter()
    manager = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 19:00:00"))

    cabinet = _cabinet_decisions()
    cabinet_bytes = adapter.export_to_bytes(
        cabinet,
        "РС УрФО Империал",
        source_name="Перемещение 01.09.xls",
    ).getvalue()
    cabinet_meta = build_order_run_meta(
        original_filename="Перемещение 01.09.xls",
        doc_type=DOC_TYPE_CABINET,
        decisions=cabinet,
        excel_bytes=cabinet_bytes,
        customer_name="РС УрФО Империал",
        timestamp="2026-08-30 19:00:01",
        order_id="scan-cab",
    )
    manager.save_run(cabinet_meta, cabinet_bytes)

    soft = _soft_decisions()
    soft_bytes = adapter.export_to_bytes(
        soft,
        "Склад мягкой мебели",
        source_name="order_transfering_01_09_bed.xls",
    ).getvalue()
    soft_meta = build_order_run_meta(
        original_filename="order_transfering_01_09_bed.xls",
        doc_type=DOC_TYPE_SOFT,
        decisions=soft,
        excel_bytes=soft_bytes,
        customer_name="Склад мягкой мебели",
        timestamp="2026-08-30 19:00:02",
        order_id="scan-soft",
    )
    manager.save_run(soft_meta, soft_bytes)

    scanned = "1234567890123"
    updated_cabinet = adapter.export_to_bytes(
        cabinet,
        "РС УрФО Империал",
        source_name="Перемещение 01.09.xls",
        overrides={2: scanned},
    ).getvalue()
    manager.update_excel(cabinet_meta, updated_cabinet)
    manager.update_session(
        cabinet_meta,
        {
            "operator_overrides": {2: scanned},
            "upload_name": "Перемещение 01.09.xls",
        },
    )

    cab_book = load_workbook(BytesIO(manager.read_excel_bytes(cabinet_meta)), data_only=True)
    cab_sheet = cab_book[WMS_SHEET_NAME]
    barcodes = [cab_sheet.cell(row, 3).value for row in range(2, cabinet_meta.total_rows + 2)]
    assert scanned in {str(value) for value in barcodes if value}

    soft_after = manager.read_excel_bytes(soft_meta)
    assert soft_after == soft_bytes
    session = manager.load_session(cabinet_meta)
    assert session is not None
    assert session["operator_overrides"][2] == scanned
    assert manager.load_session(soft_meta) is None


def test_other_date_isolated(tmp_path: Path) -> None:
    manager = HistoryManager(root=tmp_path)
    manager.save_run(_meta(timestamp="2026-08-29 08:00:00", order_id="old"), b"old")
    manager.save_run(_meta(timestamp="2026-08-30 08:00:00", order_id="new"), b"new")
    assert [item.order_id for item in manager.get_shift_history("2026-08-29")] == ["old"]
    last = manager.get_last_run("2026-08-30")
    assert last is not None
    assert last.order_id == "new"


def test_empty_shift_zip_and_missing_last(tmp_path: Path) -> None:
    manager = HistoryManager(root=tmp_path, clock=lambda: _clock("2026-08-30 10:00:00"))
    assert manager.get_last_run() is None
    assert manager.get_shift_history() == []
    raw = manager.create_shift_zip()
    with ZipFile(BytesIO(raw)) as archive:
        assert archive.namelist() == []
