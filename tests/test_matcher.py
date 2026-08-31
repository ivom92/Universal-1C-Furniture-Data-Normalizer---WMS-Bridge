"""Tests for FAISS vector store and hybrid matcher (Sprint 3)."""

from __future__ import annotations

import shutil
import tempfile
import time
from pathlib import Path
from typing import Optional
from unittest.mock import MagicMock

import faiss
import numpy as np

import pytest
from openpyxl import load_workbook

from src.adapters.wms_excel_adapter import WMSExcelAdapter
from src.matcher.feature_extractor import FeatureExtractor
from src.matcher.hybrid_matcher import HybridMatcher
from src.matcher.llm_resolver import LLMResolver
from src.matcher.vector_store import CatalogVectorStore, FAISS_AVAILABLE, NumpyVectorEngine
from src.models import CatalogEntity, LLMResolutionResponse, RawOrderBlock
from src.utils.logger import console


def _catalog_entity(
    *,
    nomenclature: str,
    nomenclature_code: str,
    packaging: Optional[str] = None,
    module: Optional[str] = None,
    label_model: Optional[str] = None,
    filling: Optional[str] = None,
    color: Optional[str] = None,
    barcode: Optional[str] = None,
) -> CatalogEntity:
    return CatalogEntity.model_validate(
        {
            "Номенклатура": nomenclature,
            "НоменклатураКод": nomenclature_code,
            "Упаковка": packaging,
            "Модуль": module,
            "ЭтикеткаМодель": label_model,
            "Начинка": filling,
            "Цвет": color,
            "Штрихкод": barcode,
        }
    )


def _raw_block(
    *,
    client_description: str,
    factory_alias: str = "",
    item_type: str = "Пачка",
    line_number: int = 1,
) -> RawOrderBlock:
    return RawOrderBlock(
        line_number=line_number,
        client_description=client_description,
        item_type=item_type,
        quantity=1,
        factory_alias=factory_alias or client_description,
        order_service_line="Продажи оптовые УРП_ test",
        excel_row_start=line_number,
    )


class FakeVectorStore:
    """Deterministic vector store stub for hard-filter unit tests."""

    def __init__(
        self,
        hits: list[tuple[CatalogEntity, float]],
        catalog: list[CatalogEntity] | None = None,
    ) -> None:
        self._hits = hits
        self._catalog = catalog if catalog is not None else []

    @property
    def catalog(self) -> list[CatalogEntity]:
        return list(self._catalog)

    def search(self, query_text: str, top_k: int = 20) -> list[tuple[CatalogEntity, float]]:
        return self._hits[:top_k]


class TestCatalogVectorStore:
    def test_vector_store_caching(self, tmp_path) -> None:
        tiny_catalog = [
            _catalog_entity(
                nomenclature="Фасад д/выдвижного ящика Плано 116х596х16 Фон милк упаковка 1/1",
                nomenclature_code="00000097658",
                packaging="1/1",
                label_model="Плано",
                module="116х596х16",
                barcode="2006000045445",
            ),
            _catalog_entity(
                nomenclature="Нео корпус 2/2 белый",
                nomenclature_code="00000010001",
                packaging="2/2",
                label_model="Нео",
            ),
        ]

        cache_dir = tmp_path / "cache"
        store = CatalogVectorStore(cache_dir=str(cache_dir))

        started = time.perf_counter()
        store.build_or_load_index(tiny_catalog)
        first_build_seconds = time.perf_counter() - started
        assert store.is_ready
        assert len(store.catalog) == 2

        hits_before = store.search("Фасад Плано 116х596 1/1", top_k=2)
        assert hits_before
        assert hits_before[0][0].nomenclature_code == "00000097658"

        reloaded = CatalogVectorStore(cache_dir=str(cache_dir))
        started = time.perf_counter()
        reloaded.build_or_load_index(tiny_catalog)
        reload_seconds = time.perf_counter() - started

        assert reloaded.is_ready
        assert reload_seconds < max(first_build_seconds, 1.0)
        hits_after = reloaded.search("Фасад Плано 116х596 1/1", top_k=2)
        assert hits_after[0][0].nomenclature_code == hits_before[0][0].nomenclature_code
        assert hits_after[0][1] == pytest.approx(hits_before[0][1], rel=1e-5)

    def test_faiss_cyrillic_path_io(self) -> None:
        root = Path(tempfile.mkdtemp(prefix="faiss_cyr_"))
        cache_dir = root / "Тестовый_Каталог_Егор" / ".cache"
        try:
            dim = 8
            vectors = np.array(
                [
                    [1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
                    [0.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
                ],
                dtype=np.float32,
            )
            faiss.normalize_L2(vectors)
            index = faiss.IndexFlatIP(dim)
            index.add(vectors)

            catalog = [
                _catalog_entity(
                    nomenclature="Фасад Плано 1/1",
                    nomenclature_code="00000097658",
                    packaging="1/1",
                    barcode="2006000045445",
                ),
                _catalog_entity(
                    nomenclature="Нео корпус 2/2",
                    nomenclature_code="00000010001",
                    packaging="2/2",
                ),
            ]

            store = CatalogVectorStore(cache_dir=str(cache_dir))
            store._index = index
            store._catalog = catalog
            saved = store.save()
            assert saved is True
            assert (cache_dir / "catalog_faiss.index").is_file()
            assert (cache_dir / "catalog_meta.pkl").is_file()

            reloaded = CatalogVectorStore(cache_dir=str(cache_dir))
            loaded = reloaded.load()
            assert loaded is True
            assert reloaded._index is not None
            assert reloaded._index.d == dim
            assert reloaded._index.ntotal == index.ntotal
            restored = np.vstack(
                [reloaded._index.reconstruct(i) for i in range(reloaded._index.ntotal)]
            )
            np.testing.assert_allclose(restored, vectors, rtol=1e-5, atol=1e-6)
            assert [item.nomenclature_code for item in reloaded.catalog] == [
                "00000097658",
                "00000010001",
            ]
            assert reloaded.catalog[0].barcode == "2006000045445"
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @pytest.mark.skipif(not FAISS_AVAILABLE, reason="FAISS not installed")
    def test_numpy_fallback_equivalent_to_faiss(self, tmp_path, monkeypatch) -> None:
        tiny_catalog = [
            _catalog_entity(
                nomenclature="Фасад д/выдвижного ящика Плано 116х596х16 Фон милк упаковка 1/1",
                nomenclature_code="00000097658",
                packaging="1/1",
                label_model="Плано",
                module="116х596х16",
                barcode="2006000045445",
            ),
            _catalog_entity(
                nomenclature="Нео корпус 2/2 белый",
                nomenclature_code="00000010001",
                packaging="2/2",
                label_model="Нео",
            ),
            _catalog_entity(
                nomenclature="Сканди стол 1400 белый",
                nomenclature_code="00000020002",
                packaging="1/1",
                label_model="Сканди",
            ),
        ]
        cache_dir = tmp_path / "cache"
        faiss_store = CatalogVectorStore(cache_dir=str(cache_dir))
        faiss_store.build_or_load_index(tiny_catalog)

        queries = [
            "Фасад Плано 116х596 1/1",
            "Нео корпус 2/2",
            "Сканди стол",
        ]
        faiss_results = {
            query: faiss_store.search(query, top_k=3) for query in queries
        }

        monkeypatch.setattr("src.matcher.vector_store.FAISS_AVAILABLE", False)
        numpy_store = CatalogVectorStore(cache_dir=str(cache_dir))
        numpy_store.build_or_load_index(tiny_catalog)

        assert numpy_store.engine_name == "NumPy Vector Engine"
        for query, expected in faiss_results.items():
            actual = numpy_store.search(query, top_k=3)
            assert actual[0][0].nomenclature_code == expected[0][0].nomenclature_code
            assert actual[0][1] == pytest.approx(expected[0][1], rel=1e-5, abs=1e-6)

    def test_numpy_engine_search_matches_dot_product(self) -> None:
        vectors = np.array(
            [
                [1.0, 0.0, 0.0],
                [0.0, 1.0, 0.0],
                [0.70710677, 0.70710677, 0.0],
            ],
            dtype=np.float32,
        )
        engine = NumpyVectorEngine(vectors)
        query = np.array([[0.70710677, 0.70710677, 0.0]], dtype=np.float32)
        scores, indices = engine.search(query, top_k=2)
        assert indices[0][0] == 2
        assert scores[0][0] == pytest.approx(1.0, rel=1e-5)


class TestHybridHardFilters:
    def test_hard_filter_packaging_isolation(self, feature_extractor: FeatureExtractor) -> None:
        same_name = "Фасад д/выдвижного ящика Плано 116х596х16 Фон милк упаковка"
        entity_1_2 = _catalog_entity(
            nomenclature=f"{same_name} 1/2",
            nomenclature_code="00000090001",
            packaging="1/2",
            module="116х596х16",
            label_model="Плано",
        )
        entity_2_2 = _catalog_entity(
            nomenclature=f"{same_name} 2/2",
            nomenclature_code="00000090002",
            packaging="2/2",
            module="116х596х16",
            label_model="Плано",
        )

        matcher = HybridMatcher(
            FakeVectorStore([(entity_1_2, 0.99), (entity_2_2, 0.95)]),
            feature_extractor,
        )
        block = _raw_block(
            client_description="Пакет фасады Плано 116х596 2/2",
            factory_alias="Пакет фасады Плано 116х596 2/2",
        )

        decision = matcher.match_block(block)

        assert decision.status in {"MATCHED_AUTO", "NEEDS_LLM"}
        passed_packaging = {
            candidate.catalog_entity.packaging
            for candidate in decision.candidates
            if candidate.hard_filter_passed
        }
        assert "1/2" not in passed_packaging
        if decision.matched_entity is not None:
            assert decision.matched_entity.packaging == "2/2"

    def test_package_ratio_hard_barrier(self, feature_extractor: FeatureExtractor) -> None:
        mismatched = [
            _catalog_entity(
                nomenclature="Ящик с доводчиком Н86 400 Ун1/1",
                nomenclature_code="00000081001",
                packaging=None,
            ),
            _catalog_entity(
                nomenclature="Кухня Равенна Н20 карго (корпус) белый упаковка 1/1",
                nomenclature_code="00000081002",
                packaging="1/1",
                label_model="Равенна",
            ),
            _catalog_entity(
                nomenclature="Кухня Равенна Н20 карго (корпус) белый упаковка 2/3",
                nomenclature_code="00000081003",
                packaging="2/3",
                label_model="Равенна",
            ),
        ]
        hits = [(entity, 0.99 - index * 0.01) for index, entity in enumerate(mismatched)]
        matcher = HybridMatcher(
            FakeVectorStore(hits, catalog=mismatched),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(
                client_description="Кухня Равенна Н20 карго (корпус) 1/3",
                factory_alias="IMP к Равенна Н20 карго (корпус) 1/3",
            )
        )
        assert decision.candidates
        for candidate in decision.candidates:
            assert candidate.hard_filter_passed is False
            assert candidate.penalty_reason == "Package ratio mismatch"
        if decision.matched_entity is not None:
            assert decision.matched_entity.packaging == "1/3"

    def test_corpus_vs_drawer_isolation(self, feature_extractor: FeatureExtractor) -> None:
        drawer = _catalog_entity(
            nomenclature="Ящик с доводчиком Н86 400 Белый упаковка 1/3",
            nomenclature_code="00000082001",
            packaging="1/3",
            color="Белый",
            barcode="2006000082001",
        )
        corpus = _catalog_entity(
            nomenclature="Кухня Равенна Н20 карго (корпус) Белый упаковка 1/3",
            nomenclature_code="00000082002",
            packaging="1/3",
            label_model="Равенна",
            color="Белый",
            barcode="2006000082002",
        )
        matcher = HybridMatcher(
            FakeVectorStore(
                [(drawer, 0.99), (corpus, 0.88)],
                catalog=[drawer, corpus],
            ),
            feature_extractor,
        )
        block = _raw_block(
            client_description="Кухня Равенна Н20 карго (корпус) Белый упаковка 1/3 без цвета",
            factory_alias="IMP к Равенна Роял Н20 карго с доводчиком = грей",
        )
        features = feature_extractor.extract_features(block)
        scored = matcher._score_candidates(block, features, [(drawer, 0.99), (corpus, 0.88)])
        drawer_candidate = next(
            candidate
            for candidate in scored
            if candidate.catalog_entity.nomenclature_code == "00000082001"
        )
        assert drawer_candidate.hard_filter_passed is False
        assert drawer_candidate.penalty_reason == "Corpus vs drawer/slide isolation"
        corpus_candidate = next(
            candidate
            for candidate in scored
            if candidate.catalog_entity.nomenclature_code == "00000082002"
        )
        assert corpus_candidate.hard_filter_passed is True

        decision = matcher.match_block(block)
        assert decision.matched_entity is not None
        assert "Ящик с доводчиком" not in decision.matched_entity.nomenclature
        assert decision.matched_entity.nomenclature_code == "00000082002"

    def test_no_barcode_status_badge(self, feature_extractor: FeatureExtractor) -> None:
        from src.utils.reporter import get_status_badge

        entity = _catalog_entity(
            nomenclature="Планка 1516 белая упаковка Ун1/1",
            nomenclature_code="00000015160",
            packaging="Ун1/1",
            barcode=None,
        )
        matcher = HybridMatcher(
            FakeVectorStore([(entity, 0.96)], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(
                client_description="Планка 1516 белая Ун1/1",
                factory_alias="Планка 1516 белая",
                item_type="Фурнитура",
            )
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.match_method in {"AUTO_NO_BARCODE", "MATCHED_AUTO_NO_BARCODE"}
        assert decision.matched_entity is not None
        assert decision.matched_entity.barcode is None
        assert get_status_badge(decision) == "🟢 Авто (без ШК)"

    def test_exact_dimension_matching(self, feature_extractor: FeatureExtractor) -> None:
        entity_116 = _catalog_entity(
            nomenclature="Фасад Плано 116х596х16 1/1",
            nomenclature_code="00000097658",
            packaging="1/1",
            module="116х596х16",
            label_model="Плано",
            barcode="2006000045445",
        )
        entity_140 = _catalog_entity(
            nomenclature="Фасад Плано 140х596х16 1/1",
            nomenclature_code="00000097659",
            packaging="1/1",
            module="140х596х16",
            label_model="Плано",
            barcode="2006000045446",
        )

        matcher = HybridMatcher(
            FakeVectorStore([(entity_140, 0.99), (entity_116, 0.97)]),
            feature_extractor,
        )
        block = _raw_block(
            client_description="Фасад 116 x 596 x 16 Плано 1/1",
            factory_alias="Фасад 116х596 1/1",
        )

        decision = matcher.match_block(block)

        passed_modules = [
            candidate.catalog_entity.module
            for candidate in decision.candidates
            if candidate.hard_filter_passed
        ]
        assert all("140х596" not in (module or "") for module in passed_modules)
        if decision.matched_entity is not None:
            assert "116х596" in (decision.matched_entity.module or "")


class TestHybridAutoMatch:
    def test_auto_match_high_confidence(self, hybrid_matcher: HybridMatcher) -> None:
        block = _raw_block(
            client_description="Фасад 116 x 596 x 16 Плано 1/1",
            factory_alias="Фасад д/выдвижного ящика Плано 116х596х16 Фон милк упаковка 1/1",
        )

        decision = hybrid_matcher.match_block(block)

        assert decision.status == "MATCHED_AUTO"
        assert decision.confidence_score >= 0.90
        assert decision.matched_entity is not None
        assert decision.matched_entity.barcode == "2006000045445"
        assert isinstance(decision.matched_entity.nomenclature_code, str)
        assert isinstance(decision.matched_entity.barcode, str)


class TestHybridIntegration:
    def test_integration_ruban_order(
        self,
        hybrid_matcher: HybridMatcher,
        order_ruban_parsed,
    ) -> None:
        blocks = order_ruban_parsed.blocks
        customer_name = order_ruban_parsed.customer_name

        matched_items = hybrid_matcher.match_order(blocks, customer_name)

        assert len(matched_items) == 55
        assert len(blocks) == len(matched_items)

        status_counts = {"MATCHED_AUTO": 0, "NEEDS_LLM": 0, "QUARANTINE": 0}
        for item in matched_items:
            reason = item.match_reason or "QUARANTINE"
            if reason in {
                "vector_auto",
                "AUTO_NO_BARCODE",
                "exact_article",
                "exact_article_no_barcode",
            }:
                status_counts["MATCHED_AUTO"] += 1
            elif reason == "NEEDS_LLM":
                status_counts["NEEDS_LLM"] += 1
            else:
                status_counts["QUARANTINE"] += 1

            if item.nomenclature_code is not None:
                assert isinstance(item.nomenclature_code, str)
            if item.barcode is not None:
                assert isinstance(item.barcode, str)

        total = len(matched_items)
        auto_pct = status_counts["MATCHED_AUTO"] / total * 100
        llm_pct = status_counts["NEEDS_LLM"] / total * 100
        quarantine_pct = status_counts["QUARANTINE"] / total * 100

        console.print(
            "[bold]Ruban order matching summary[/bold]: "
            f"MATCHED_AUTO={status_counts['MATCHED_AUTO']} ({auto_pct:.1f}%), "
            f"NEEDS_LLM={status_counts['NEEDS_LLM']} ({llm_pct:.1f}%), "
            f"QUARANTINE={status_counts['QUARANTINE']} ({quarantine_pct:.1f}%)"
        )

        assert status_counts["MATCHED_AUTO"] + status_counts["NEEDS_LLM"] + status_counts["QUARANTINE"] == 55
        assert status_counts["MATCHED_AUTO"] >= 40


class TestNoBarcodeIsNotQuarantine:
    def test_catalog_hit_without_ean_is_matched_auto(
        self,
        feature_extractor: FeatureExtractor,
        tmp_path,
    ) -> None:
        entity = _catalog_entity(
            nomenclature="Планка декоративная Аврора 1/1 венге",
            nomenclature_code="00000055501",
            packaging="1/1",
            label_model="Аврора",
            color="венге",
            barcode=None,
        )
        matcher = HybridMatcher(
            FakeVectorStore([(entity, 0.95)]),
            feature_extractor,
        )
        block = _raw_block(
            client_description="Планка декоративная Аврора 1/1 венге",
            factory_alias="Планка декоративная Аврора 1/1 венге",
        )

        decision = matcher.match_block(block)

        assert decision.status == "MATCHED_AUTO"
        assert decision.status != "QUARANTINE"
        assert decision.matched_entity is not None
        assert decision.matched_entity.barcode is None
        assert decision.match_method == "AUTO_NO_BARCODE"

        item = matcher.match_order([block], "Тестовый заказчик")[0]
        assert item.nomenclature == entity.nomenclature
        assert item.barcode is None
        assert item.nomenclature_code == "00000055501"

        export_path = tmp_path / "wms_no_barcode.xlsx"
        WMSExcelAdapter().export([decision], "Тестовый заказчик", export_path)
        worksheet = load_workbook(export_path).active
        assert worksheet.cell(row=2, column=2).value == entity.nomenclature
        assert worksheet.cell(row=2, column=3).value in (None, "")

    def test_llm_hit_without_ean_is_matched_llm(
        self,
        feature_extractor: FeatureExtractor,
    ) -> None:
        entity = _catalog_entity(
            nomenclature="Заглушка Аврора 1/1 венге",
            nomenclature_code="00000055502",
            packaging="1/1",
            label_model="Аврора",
            color="венге",
            barcode=None,
        )
        mock_resolver = MagicMock(spec=LLMResolver)
        mock_resolver.provider = "gemini"
        mock_resolver.resolve.return_value = LLMResolutionResponse(
            selected_nomenclature_code="00000055502",
            confidence=0.91,
            reasoning="фурнитура без заводского ШК",
        )
        matcher = HybridMatcher(
            FakeVectorStore([(entity, 0.70)]),
            feature_extractor,
            llm_resolver=mock_resolver,
        )
        block = _raw_block(
            client_description="Заглушка Аврора 1/1 венге",
            factory_alias="Заглушка Аврора 1/1 венге",
        )

        decision = matcher.match_block(block)

        assert decision.status == "MATCHED_LLM"
        assert decision.matched_entity is not None
        assert decision.matched_entity.barcode is None
        assert decision.match_method == "LLM_NO_BARCODE"


class TestFeatureBoostAurora:
    @pytest.mark.parametrize(
        ("description", "nomenclature", "packaging", "code"),
        [
            (
                "Аврора Зеркало 1/1 венге",
                "Аврора Зеркало 1/1 венге",
                "1/1",
                "00000070001",
            ),
            (
                "Аврора Кровать 140 1/2 венге",
                "Аврора Кровать 140 1/2 венге",
                "1/2",
                "00000070002",
            ),
            (
                "Аврора Кровать 140 2/2 венге",
                "Аврора Кровать 140 2/2 венге",
                "2/2",
                "00000070003",
            ),
        ],
    )
    def test_aurora_feature_boost_reaches_auto_threshold(
        self,
        feature_extractor: FeatureExtractor,
        description: str,
        nomenclature: str,
        packaging: str,
        code: str,
    ) -> None:
        entity = _catalog_entity(
            nomenclature=nomenclature,
            nomenclature_code=code,
            packaging=packaging,
            label_model="Аврора",
            color="венге",
            barcode="2006000070001",
        )
        matcher = HybridMatcher(
            FakeVectorStore([(entity, 0.78)]),
            feature_extractor,
        )
        decision = matcher.match_block(_raw_block(client_description=description))

        assert decision.status == "MATCHED_AUTO"
        assert decision.match_method == "vector_auto"
        assert decision.confidence_score >= 0.83
        assert decision.matched_entity is not None
        assert decision.matched_entity.packaging == packaging
        boost = HybridMatcher._compute_feature_boost(
            decision.extracted_features,
            entity,
        )
        assert boost == pytest.approx(0.08)
        assert decision.confidence_score == pytest.approx(min(1.0, 0.78 + boost))


class TestLexicalArticleSearch:
    def test_planka_1516_exact_article(self, feature_extractor: FeatureExtractor) -> None:
        entity = _catalog_entity(
            nomenclature="Планка 1516 угловая 38мм 1/1",
            nomenclature_code="00000015160",
            packaging="1/1",
            barcode="2006000015160",
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(client_description="Планка 1516 угловая 38мм 1/1")
        )

        assert decision.status == "MATCHED_AUTO"
        assert decision.confidence_score == 1.0
        assert decision.match_method == "exact_article"
        assert decision.matched_entity is not None
        assert "1516" in decision.matched_entity.nomenclature

    def test_lamp_dotted_article(self, feature_extractor: FeatureExtractor) -> None:
        entity = _catalog_entity(
            nomenclature="Светильник LED 04.002.20.312 1/1",
            nomenclature_code="00000004002",
            packaging="1/1",
            barcode="2006000004002",
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(client_description="Светильник 04.002.20.312")
        )

        assert decision.status == "MATCHED_AUTO"
        assert decision.confidence_score == 1.0
        assert decision.matched_entity is not None
        assert "04.002.20.312" in decision.matched_entity.nomenclature

    def test_chicago_multi_size_matches_1600(self, feature_extractor: FeatureExtractor) -> None:
        entity_1600 = _catalog_entity(
            nomenclature="Кровать Чикаго Нео 1600 1/1",
            nomenclature_code="00000016001",
            packaging="1/1",
            label_model="Чикаго",
            barcode="2006000016001",
        )
        entity_1800 = _catalog_entity(
            nomenclature="Кровать Чикаго Нео 1800 1/1",
            nomenclature_code="00000018001",
            packaging="1/1",
            label_model="Чикаго",
            barcode="2006000018001",
        )
        matcher = HybridMatcher(
            FakeVectorStore([(entity_1800, 0.96), (entity_1600, 0.94)]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(client_description="Чикаго Нео 160/140/120 1/1")
        )

        assert decision.extracted_features.alternative_widths == [1600, 1400, 1200]
        assert decision.extracted_features.package_ratio == "1/1"
        passed_codes = {
            candidate.catalog_entity.nomenclature_code
            for candidate in decision.candidates
            if candidate.hard_filter_passed
        }
        assert "00000016001" in passed_codes
        assert "00000018001" not in passed_codes
        assert decision.status == "MATCHED_AUTO"
        assert decision.matched_entity is not None
        assert "1600" in decision.matched_entity.nomenclature

    def test_corner_unique_hardware_phrase(self, feature_extractor: FeatureExtractor) -> None:
        entity = _catalog_entity(
            nomenclature="Корнер универсальный 1/1",
            nomenclature_code="00000017540",
            packaging="1/1",
            barcode="2006000017540",
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(_raw_block(client_description="Корнер 1/1"))
        assert decision.status == "MATCHED_AUTO"
        assert decision.confidence_score == 1.0
        assert decision.matched_entity is not None
        assert "Корнер" in decision.matched_entity.nomenclature

    def test_quarantine_reason_when_article_missing(
        self,
        feature_extractor: FeatureExtractor,
    ) -> None:
        entity = _catalog_entity(
            nomenclature="Планка 1516 угловая 38мм 1/1",
            nomenclature_code="00000015160",
            packaging="1/1",
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(client_description="Планка 9999 угловая 38мм 1/1")
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.match_method == "AUTO_NO_BARCODE"
        assert decision.matched_entity is None


class TestHardwareContourAndAutoPromotion:
    def test_corner_plug_auto_no_barcode(self, feature_extractor: FeatureExtractor) -> None:
        entity = _catalog_entity(
            nomenclature="Корнер заглушка ЛВ 1/1 (476) белый",
            nomenclature_code="00000017476",
            packaging="1/1",
            color="белый",
            barcode=None,
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(client_description="Корнер заглушка ЛВ 1/1 (476) белый")
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.match_method == "AUTO_NO_BARCODE"
        assert decision.matched_entity is not None
        assert decision.matched_entity.barcode is None

    def test_planka_1516_finds_v8(self, feature_extractor: FeatureExtractor) -> None:
        entity = _catalog_entity(
            nomenclature="Планка 1516 угловая 38мм 1/1",
            nomenclature_code="00000015160",
            packaging="1/1",
            barcode="2006000015160",
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(client_description="Планка 1516 угловая 38мм 1/1")
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.matched_entity is not None
        assert "1516" in decision.matched_entity.nomenclature

    def test_plinth_angle_gray(self, feature_extractor: FeatureExtractor) -> None:
        entity = _catalog_entity(
            nomenclature="Угол 135 гр для цоколя 100 мм грей",
            nomenclature_code="00000018135",
            packaging="1/1",
            color="грей",
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(client_description="Угол 135 гр для цоколя 100 мм грей")
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.matched_entity is not None
        assert "цокол" in decision.matched_entity.nomenclature.lower()

    def test_rus_plinth_wotan(self, feature_extractor: FeatureExtractor) -> None:
        entity = _catalog_entity(
            nomenclature="Плинтус д/столешн. RUS 3 м 1/1 Дуб вотан (RUS-17)",
            nomenclature_code="00000017170",
            packaging="1/1",
            color="Дуб вотан",
            barcode="2006000017170",
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(
                client_description="Плинтус д/столешн. RUS 3 м 1/1 Дуб вотан (RUS-17)"
            )
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.matched_entity is not None
        assert "RUS" in decision.matched_entity.nomenclature

    def test_lazio_scandi_high_confidence_auto(
        self,
        feature_extractor: FeatureExtractor,
    ) -> None:
        entity = _catalog_entity(
            nomenclature="Система Лацио Сканди Витрина 1д корпус упаковка 1/4 Вотан",
            nomenclature_code="00000018801",
            packaging="1/4",
            label_model="Лацио Сканди",
            module="Витрина 1д",
            color="Вотан",
            barcode="2006000018801",
        )
        matcher = HybridMatcher(
            FakeVectorStore([(entity, 0.70)]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(
                client_description="Лацио Сканди Витрина 1д (корпус) 1/4 Вотан"
            )
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.match_method == "vector_auto"
        assert decision.confidence_score >= 0.88
        assert decision.matched_entity is not None
        assert decision.matched_entity.packaging == "1/4"


class TestFacadeFinishAndStartSlides:
    def test_ravenna_fe_suffix_is_hard_filter(self, feature_extractor: FeatureExtractor) -> None:
        plain = _catalog_entity(
            nomenclature="Кухня Равенна Н60 2ящ (корпус) Белый упаковка 1/3",
            nomenclature_code="00000051001",
            packaging="1/3",
            label_model="Кухня Равенна",
            module="Н60 2ящ",
            color="Белый",
            barcode="2006000051001",
        )
        enamel = _catalog_entity(
            nomenclature="Кухня Равенна Н60 2ящ (FE) (корпус) Белый упаковка 1/3",
            nomenclature_code="00000051002",
            packaging="1/3",
            label_model="Кухня Равенна",
            module="Н60 2ящ (FE)",
            color="Белый",
            barcode="2006000051002",
        )
        matcher = HybridMatcher(
            FakeVectorStore([(enamel, 0.91), (plain, 0.90)], catalog=[plain, enamel]),
            feature_extractor,
        )
        without_fe = matcher.match_block(
            _raw_block(
                client_description="Кухня Равенна Н60 2ящ (корпус) Белый упаковка 1/3",
                factory_alias="IMP к Равенна Фиеста Н60 2ящ = грин грей",
            )
        )
        assert without_fe.status == "MATCHED_AUTO"
        assert without_fe.matched_entity is not None
        assert without_fe.matched_entity.nomenclature_code == "00000051001"

        with_fe = matcher.match_block(
            _raw_block(
                client_description="Кухня Равенна Н60 2ящ (FE) (корпус) Белый упаковка 1/3",
                factory_alias="IMP к Равенна Роял Н60 2ящ (FE) = скай",
            )
        )
        assert with_fe.status in {"MATCHED_AUTO", "NEEDS_LLM"}
        if with_fe.status == "MATCHED_AUTO":
            assert with_fe.matched_entity is not None
            assert "(FE)" in with_fe.matched_entity.nomenclature

    def test_start_drawer_slide_matches_catalog(self, feature_extractor: FeatureExtractor) -> None:
        entity = _catalog_entity(
            nomenclature="Ящик с доводчиком Н86 400 Белый упаковка Ун1/1",
            nomenclature_code="00000038401",
            packaging="Ун1/1",
            color="Белый",
            barcode="2006000038401",
        )
        matcher = HybridMatcher(
            FakeVectorStore([(entity, 0.72)], catalog=[entity]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(
                client_description="Ящик с доводчиком Н86 400 белый СТАРТ без цвета",
                factory_alias="IMP к Равенна Спарк Н80 2ящ (SB) = белый/фон фреско",
                item_type="Фурнитура",
            )
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.matched_entity is not None
        assert decision.matched_entity.nomenclature_code == "00000038401"

    def test_glass_shelf_supports_are_auto_no_barcode(
        self,
        feature_extractor: FeatureExtractor,
    ) -> None:
        kitchen_glass = _catalog_entity(
            nomenclature="Кухня Равенна Полка стеклянная 565х255 1/1",
            nomenclature_code="00000056501",
            packaging="1/1",
            barcode="2006000056501",
        )
        matcher = HybridMatcher(
            FakeVectorStore([(kitchen_glass, 0.91)], catalog=[kitchen_glass]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(
                client_description=(
                    "Фурнитура Полка стеклянная 30/40/50/60/80 "
                    "(полкодержатели 8 шт) без цвета"
                ),
                factory_alias="Фурнитура Полка стеклянная 30/40/50/60/80 (полкодержатели 8 шт)",
                item_type="Фурнитура",
            )
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.match_method == "AUTO_NO_BARCODE"
        assert decision.matched_entity is None
        assert decision.status != "QUARANTINE"

    def test_cyrillic_vs_latin_dimensions_exact_match(
        self,
        feature_extractor: FeatureExtractor,
    ) -> None:
        catalog_row = _catalog_entity(
            nomenclature="Кухня Равенна Полка стеклянная 565x255 упаковка 1/1",
            nomenclature_code="00000025467",
            packaging="1/1",
            barcode="4603734801972",
        )
        matcher = HybridMatcher(
            FakeVectorStore([], catalog=[catalog_row]),
            feature_extractor,
        )
        decision = matcher.match_block(
            _raw_block(
                client_description=(
                    "Кухня Равенна полка стеклянная 60 (2 шт 5мм) 565х255 упаковка 1/1"
                ),
                factory_alias="Кухня Равенна полка стеклянная 565х255 1/1",
                item_type="Пачка",
            )
        )
        assert decision.status == "MATCHED_AUTO"
        assert decision.matched_entity is not None
        assert decision.matched_entity.nomenclature_code == "00000025467"
        assert decision.matched_entity.barcode == "4603734801972"

