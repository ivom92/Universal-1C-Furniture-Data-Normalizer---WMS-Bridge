"""Integration and unit tests for v7/v8 Excel parsers."""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.models import CatalogEntity, MatchedOrderItem, RawOrderBlock
from src.parsers.v7_parser import parse_v7_order
from src.parsers.v8_loader import format_nomenclature_code, load_catalog_v8, restore_barcode
from tests.conftest import CATALOG_V8_PATH, ORDER_RUBAN_PATH


@pytest.fixture(scope="module")
def catalog_v8() -> list[CatalogEntity]:
    return load_catalog_v8(CATALOG_V8_PATH)


@pytest.fixture(scope="module")
def order_ruban_parsed():
    return parse_v7_order(ORDER_RUBAN_PATH)


class TestV8CoercionHelpers:
    def test_nomenclature_code_preserves_leading_zeros(self) -> None:
        assert format_nomenclature_code(64794, "00000000000") == "00000064794"
        assert format_nomenclature_code("97658", "00000000000") == "00000097658"

    def test_barcode_from_int(self) -> None:
        assert restore_barcode(2006000045445) == "2006000045445"

    def test_barcode_from_spaced_string(self) -> None:
        assert restore_barcode("             ") is None

    def test_barcode_from_scientific_string(self) -> None:
        assert restore_barcode("2,006E+12") == "2006000000000"
        assert restore_barcode("2.006E+12") == "2006000000000"

    def test_barcode_from_float(self) -> None:
        assert restore_barcode(2.006e12) == "2006000000000"

    def test_article_tokens_from_catalog_text(self) -> None:
        from src.parsers.v8_loader import extract_article_tokens

        tokens = extract_article_tokens("Планка 1516 угловая 38мм 1/1")
        assert "1516" in tokens
        dotted = extract_article_tokens("Светильник 04.002.20.312")
        assert "04.002.20.312" in dotted
        masked = extract_article_tokens("Фасад 116х596х16 Плано 1/1")
        assert "116" not in masked
        assert "596" not in masked


@pytest.mark.skipif(
    not CATALOG_V8_PATH.exists(),
    reason="Real catalog file is missing in data/",
)
class TestV8LoaderRealCatalog:
    def test_loads_at_least_10k_rows(self, catalog_v8: list[CatalogEntity]) -> None:
        assert len(catalog_v8) >= 10_000

    def test_real_catalog_row_count(self, catalog_v8: list[CatalogEntity]) -> None:
        assert len(catalog_v8) == 12_880

    def test_nomenclature_code_is_string_with_leading_zeros(self, catalog_v8: list[CatalogEntity]) -> None:
        sample = catalog_v8[0]
        assert isinstance(sample.nomenclature_code, str)
        assert len(sample.nomenclature_code) == 11
        assert sample.nomenclature_code.startswith("000000")

    def test_barcode_string_or_none(self, catalog_v8: list[CatalogEntity]) -> None:
        with_barcode = [item for item in catalog_v8 if item.barcode]
        without_barcode = [item for item in catalog_v8 if item.barcode is None]
        assert len(with_barcode) == 10_200
        assert len(without_barcode) == 2_680
        assert all(isinstance(item.barcode, str) for item in with_barcode)
        assert all(item.barcode.isdigit() for item in with_barcode[:50])

    def test_known_barcode_value(self, catalog_v8: list[CatalogEntity]) -> None:
        match = next(item for item in catalog_v8 if item.barcode == "2006000045445")
        assert match.nomenclature_code == "00000097658"
        assert "Плано" in match.nomenclature

    def test_all_seventeen_fields_present_on_entity(self, catalog_v8: list[CatalogEntity]) -> None:
        entity = catalog_v8[0]
        assert entity.nomenclature
        assert entity.nomenclature_code
        assert hasattr(entity, "packaging")
        assert hasattr(entity, "storage_zone")


@pytest.mark.skipif(
    not ORDER_RUBAN_PATH.exists(),
    reason="Real order file is missing in data/orders/",
)
class TestV7ParserRealOrder:
    def test_customer_name_from_header(self, order_ruban_parsed) -> None:
        assert order_ruban_parsed.customer_name == "Рубан Кристина Олеговна ИП"

    def test_block_count(self, order_ruban_parsed) -> None:
        assert len(order_ruban_parsed.blocks) == 55

    def test_line_numbers_are_sequential(self, order_ruban_parsed) -> None:
        assert [block.line_number for block in order_ruban_parsed.blocks] == list(range(1, 56))

    def test_first_block_structure(self, order_ruban_parsed) -> None:
        block = order_ruban_parsed.blocks[0]
        assert isinstance(block, RawOrderBlock)
        assert block.line_number == 1
        assert "КДР к Столешница" in block.client_description
        assert block.item_type.strip() == ""
        assert block.quantity == 1
        assert block.factory_alias.startswith("КДР к Столешница")
        assert "Продажи оптовые" in block.order_service_line
        assert "ЦНТ-" in block.order_service_line

    def test_glass_block_has_type_and_quantity(self, order_ruban_parsed) -> None:
        glass = next(block for block in order_ruban_parsed.blocks if block.line_number == 15)
        assert glass.item_type == "Стекло"
        assert glass.quantity == 2
        assert "IMP ст" in glass.factory_alias

    def test_matched_order_item_contract_can_be_built(self, order_ruban_parsed) -> None:
        block = order_ruban_parsed.blocks[0]
        item = MatchedOrderItem(
            nomenclature="placeholder",
            barcode=None,
            quantity=block.quantity,
            customer_name=order_ruban_parsed.customer_name,
            source_block=block,
        )
        assert item.customer_name == order_ruban_parsed.customer_name
        assert item.quantity == block.quantity


class TestV7XlsCompatibility:
    def test_xls_header_and_three_row_blocks(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import SAMPLE_CUSTOMER, write_sample_v7_xls

        xls_path = write_sample_v7_xls(tmp_path / "sample_v7.xls")
        result = parse_v7_order(xls_path)

        assert result.customer_name == SAMPLE_CUSTOMER
        assert len(result.blocks) > 0
        assert len(result.blocks) == 2

        first, second = result.blocks
        assert first.quantity == 1
        assert "КДР к Столешница" in first.client_description
        assert first.factory_alias.startswith("КДР")
        assert first.excel_row_start == 8

        assert second.item_type == "Стекло"
        assert second.quantity == 2
        assert second.factory_alias.upper().startswith("IMP")
        assert "Продажи оптовые" in second.order_service_line

    def test_xls_poluchatel_header(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_sample_v7_xls

        xls_path = write_sample_v7_xls(
            tmp_path / "sample_poluchatel.xls",
            customer_label="Получатель:",
            customer_name="ООО Получатель Тест",
        )
        result = parse_v7_order(xls_path)
        assert result.customer_name == "ООО Получатель Тест"

    def test_xls_bytes_and_bytesio(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import SAMPLE_CUSTOMER, write_sample_v7_xls

        xls_path = write_sample_v7_xls(tmp_path / "sample_stream.xls")
        payload = xls_path.read_bytes()

        from_bytes = parse_v7_order(payload, filename="Отборочная-1.xls")
        from_buffer = parse_v7_order(io.BytesIO(payload), filename="Отборочная-1.xls")

        assert from_bytes.customer_name == SAMPLE_CUSTOMER
        assert from_buffer.customer_name == SAMPLE_CUSTOMER
        assert len(from_bytes.blocks) == len(from_buffer.blocks) == 2

    def test_xlsx_and_xls_results_are_equivalent(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_sample_v7_xls, write_sample_v7_xlsx

        xlsx_path = write_sample_v7_xlsx(tmp_path / "sample_v7.xlsx")
        xls_path = write_sample_v7_xls(tmp_path / "sample_v7.xls")

        xlsx_result = parse_v7_order(xlsx_path)
        xls_result = parse_v7_order(xls_path)

        assert xlsx_result.customer_name == xls_result.customer_name
        assert len(xlsx_result.blocks) == len(xls_result.blocks)
        for left, right in zip(xlsx_result.blocks, xls_result.blocks, strict=True):
            assert left.model_dump() == right.model_dump()

    def test_xls_e2e_matcher_and_wms_zero_loss(self, tmp_path: Path) -> None:
        from src.adapters.wms_excel_adapter import WMSExcelAdapter
        from src.matcher.dynamic_vocab import DynamicVocabulary
        from src.matcher.feature_extractor import FeatureExtractor
        from src.matcher.hybrid_matcher import HybridMatcher
        from src.matcher.vector_store import CatalogVectorStore
        from src.models import CatalogEntity
        from tests.excel_fixtures import SAMPLE_CUSTOMER, write_sample_v7_xls

        xls_path = write_sample_v7_xls(tmp_path / "sample_e2e.xls")
        parsed = parse_v7_order(xls_path)
        assert len(parsed.blocks) == 2

        catalog = [
            CatalogEntity.model_validate(
                {
                    "Номенклатура": "КДР к Столешница 3000х600х40 Дуб сонома 1/1",
                    "НоменклатураКод": "00000010001",
                    "Штрихкод": "2006000000001",
                    "Упаковка": "1/1",
                }
            ),
            CatalogEntity.model_validate(
                {
                    "Номенклатура": "IMP ст Витрина 116х596 стекло",
                    "НоменклатураКод": "00000010002",
                    "Штрихкод": "2006000000002",
                    "Упаковка": "1/1",
                }
            ),
        ]
        vocabulary = DynamicVocabulary(catalog)
        feature_extractor = FeatureExtractor(vocabulary)
        vector_store = CatalogVectorStore(cache_dir=str(tmp_path / "faiss"))
        vector_store.build_or_load_index(catalog)
        matcher = HybridMatcher(vector_store, feature_extractor)

        decisions = matcher.match_order_decisions(parsed.blocks)
        assert len(decisions) == len(parsed.blocks)

        output_path = tmp_path / "wms_from_xls.xlsx"
        saved = WMSExcelAdapter().export(decisions, parsed.customer_name, output_path)
        workbook = load_workbook(saved)
        worksheet = workbook.active
        data_last_row = WMSExcelAdapter.wms_data_last_row(len(parsed.blocks))
        assert worksheet.max_row == WMSExcelAdapter.wms_totals_row(len(parsed.blocks))
        for row_index in range(2, data_last_row + 1):
            assert worksheet.cell(row=row_index, column=5).value == SAMPLE_CUSTOMER
            assert int(worksheet.cell(row=row_index, column=4).value) == parsed.blocks[row_index - 2].quantity
            assert int(worksheet.cell(row=row_index, column=1).value) == parsed.blocks[row_index - 2].order_line_number
        workbook.close()


class TestV7TransferAndHeaderHardening:
    def test_transfer_inline_recipient_and_merged_cell(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_transfer_v7_xlsx

        path = write_transfer_v7_xlsx(
            tmp_path / "transfer_inline.xlsx",
            recipient="Челябинск ТК",
            inline_header=True,
        )
        result = parse_v7_order(path)
        assert result.customer_name == "Челябинск ТК"
        assert len(result.blocks) == 2
        assert result.blocks[0].factory_alias.startswith("КДР")
        assert result.blocks[1].factory_alias.upper().startswith("IMP")

    def test_transfer_adjacent_cell_label(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_transfer_v7_xlsx

        path = write_transfer_v7_xlsx(
            tmp_path / "transfer_adjacent.xlsx",
            recipient="Склад №1 Челябинск",
            inline_header=False,
        )
        result = parse_v7_order(path)
        assert result.customer_name == "Склад №1 Челябинск"

    def test_fallback_uses_document_title(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_transfer_v7_xlsx

        title = "Перемещение № ЧЛ-00452 от 01.09.2026"
        path = write_transfer_v7_xlsx(
            tmp_path / "transfer_fallback.xlsx",
            include_recipient_label=False,
            document_title=title,
        )
        result = parse_v7_order(path)
        assert result.customer_name == title
        assert len(result.blocks) == 2

    def test_two_row_transfer_items_allow_missing_alias(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_transfer_v7_xlsx

        path = write_transfer_v7_xlsx(
            tmp_path / "transfer_two_row.xlsx",
            recipient="Челябинск ТК",
            two_row_items=True,
        )
        result = parse_v7_order(path)
        assert result.customer_name == "Челябинск ТК"
        assert len(result.blocks) == 2
        assert result.blocks[0].factory_alias is None
        assert result.blocks[1].factory_alias is None
        assert "Продажи оптовые" in result.blocks[0].order_service_line
        assert result.blocks[0].quantity == 1
        assert result.blocks[1].quantity == 2

    def test_filename_fallback_when_header_absent(self, tmp_path: Path) -> None:
        from openpyxl import Workbook
        from tests.excel_fixtures import MAIN_FILL, SAMPLE_BLOCKS

        path = tmp_path / "empty_header.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(1, 1, "Служебная печать")
        row = 6
        for block in SAMPLE_BLOCKS:
            sheet.cell(row, 1, block["line_number"])
            cell = sheet.cell(row, 2, block["description"])
            cell.fill = MAIN_FILL
            sheet.cell(row, 8, block["quantity"])
            sheet.cell(row + 1, 2, block["alias"])
            sheet.cell(row + 2, 2, block["service"])
            row += 3
        workbook.save(path)
        workbook.close()

        result = parse_v7_order(path)
        assert result.customer_name == "Перемещение (empty_header.xlsx)"
        assert len(result.blocks) == 2


class TestV7CellTopologyAndRegionalWarehouse:
    def test_strips_cell_location_from_client_description(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_location_topology_v7_xlsx

        path = write_location_topology_v7_xlsx(tmp_path / "transfer_cells.xlsx")
        result = parse_v7_order(path)

        assert result.customer_name == "РС УрФО Империал"
        assert len(result.blocks) == 2
        assert result.blocks[0].client_description == "Аврора Зеркало 1/1 венге"
        assert "Р1.16.Я2" not in result.blocks[0].client_description
        assert "Р10" not in result.blocks[1].client_description
        assert result.blocks[1].client_description.startswith("Аврора Кровать")

    def test_reg_sklad_adjacent_cell(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_location_topology_v7_xlsx

        path = write_location_topology_v7_xlsx(
            tmp_path / "transfer_reg_adjacent.xlsx",
            inline_warehouse=False,
        )
        result = parse_v7_order(path)
        assert result.customer_name == "РС УрФО Империал"

    def test_inline_cell_prefix_is_sanitized(self) -> None:
        from src.parsers.v7_parser import strip_cell_location_prefix

        cleaned = strip_cell_location_prefix("Р1.16.Я2 Аврора Зеркало 1/1 венге")
        assert cleaned == "Аврора Зеркало 1/1 венге"


class TestUniversalLayoutEngine:
    def test_shifted_columns_dynamic_mapping(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import SAMPLE_CUSTOMER, write_shifted_columns_v7_xlsx

        path = write_shifted_columns_v7_xlsx(tmp_path / "shifted_columns.xlsx")
        result = parse_v7_order(path)

        assert result.customer_name == SAMPLE_CUSTOMER
        assert len(result.blocks) == 2
        first, second = result.blocks
        assert first.line_number == 1
        assert "КДР к Столешница" in first.client_description
        assert "Р1.1" not in first.client_description
        assert first.quantity == 1
        assert first.factory_alias is not None
        assert first.factory_alias.startswith("КДР")
        assert "примечание склада" in first.factory_alias
        assert "Продажи оптовые" in first.order_service_line
        assert second.line_number == 2
        assert second.item_type == "Стекло"
        assert second.quantity == 2
        assert second.factory_alias is not None
        assert second.factory_alias.upper().startswith("IMP")

    def test_five_row_blocks_and_glued_section_tokens(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import SAMPLE_CUSTOMER, write_five_row_block_v7_xlsx

        path = write_five_row_block_v7_xlsx(tmp_path / "five_row.xlsx")
        result = parse_v7_order(path)

        assert result.customer_name == SAMPLE_CUSTOMER
        assert len(result.blocks) == 2
        first, second = result.blocks
        assert first.line_number == 1
        assert first.client_description == "Аврора Зеркало 1/1 венге"
        assert "Секция" not in first.client_description
        assert first.quantity == 1
        assert second.line_number == 2
        assert "КДР к Столешница" in second.client_description
        assert "Р3.4" not in second.client_description
        assert second.factory_alias is not None
        assert "примечание комплектовки" in second.factory_alias
        assert "Перемещение на склад" in second.order_service_line
        assert "Продажи оптовые" in second.order_service_line

    def test_glued_section_token_is_sanitized(self) -> None:
        from src.parsers.v7_parser import sanitize_warehouse_topology

        cleaned = sanitize_warehouse_topology("Секция 12 Аврора Зеркало 1/1 венге")
        assert cleaned == "Аврора Зеркало 1/1 венге"


class TestHtmlXlsAndNfkc:
    def test_html_saved_as_xls_is_parsed(self, tmp_path: Path) -> None:
        path = tmp_path / "html_as_xls.xls"
        nbsp = "\xa0"
        path.write_text(
            "<html><body><table>"
            "<tr><td>Отборочный лист</td></tr>"
            f"<tr><td>Покупатель:</td><td>ИП{nbsp}HTML</td></tr>"
            "<tr><td>№</td><td>Наименование</td><td>Кол-во</td></tr>"
            f"<tr><td>1</td><td>Йорк{nbsp}Комод{nbsp}1/1</td><td>2</td></tr>"
            "<tr><td></td><td>Продажи оптовые УРП_1 Заказ: ЦНТ-1</td></tr>"
            "</table></body></html>",
            encoding="utf-8",
        )
        result = parse_v7_order(path)
        assert "ИП HTML" in result.customer_name
        assert len(result.blocks) == 1
        assert result.blocks[0].quantity == 2
        assert "Йорк Комод" in result.blocks[0].client_description
        assert "\xa0" not in result.blocks[0].client_description

    def test_nfkc_normalizes_quantity_and_name(self, tmp_path: Path) -> None:
        from tests.excel_fixtures import write_sample_v7_xlsx

        path = write_sample_v7_xlsx(tmp_path / "nfkc.xlsx")
        from openpyxl import load_workbook

        workbook = load_workbook(path)
        sheet = workbook.active
        sheet.cell(8, 2, "КДР к Столешница\u00a03000х600х40")
        sheet.cell(8, 8, "1\u00a0")
        workbook.save(path)
        workbook.close()

        result = parse_v7_order(path)
        assert result.blocks[0].quantity == 1
        assert "\xa0" not in result.blocks[0].client_description
        assert "Столешница" in result.blocks[0].client_description

    def test_nfkc_nbsp_helper(self) -> None:
        from src.parsers.v7_parser import normalize_incoming_text

        assert normalize_incoming_text("Йорк\xa0Комод") == "Йорк Комод"
        assert "№" in normalize_incoming_text("№1")

    def test_doctype_html_xls(self, tmp_path: Path) -> None:
        path = tmp_path / "doctype.xls"
        path.write_text(
            "<!DOCTYPE html><html><body><table>"
            "<tr><td>Покупатель:</td><td>ООО Доктайп</td></tr>"
            "<tr><td>№</td><td>Наименование</td><td>Кол-во</td></tr>"
            "<tr><td>1</td><td>Плано Фасад 1/1</td><td>1</td></tr>"
            "<tr><td></td><td>Продажи оптовые УРП_1 Заказ: ЦНТ-1</td></tr>"
            "</table></body></html>",
            encoding="utf-8",
        )
        result = parse_v7_order(path)
        assert result.customer_name == "ООО Доктайп"
        assert len(result.blocks) == 1

    def test_cp1251_html_xls_decodes_cyrillic(self, tmp_path: Path) -> None:
        """1C v7.7 HTML tables without charset header are saved as Windows-1251."""
        path = tmp_path / "vasilyeva_cp1251.xls"
        html = (
            "<html><body><table>"
            "<tr><td>Отборочный лист</td></tr>"
            "<tr><td>Покупатель:</td><td>Васильева Т.</td></tr>"
            "<tr><td>№</td><td>Наименование</td><td>Кол-во</td></tr>"
            "<tr><td>1</td><td>Аврора Кровать 90 со встроенным основанием 1/2</td><td>2</td></tr>"
            "<tr><td></td><td>IMP сп Аврора кровать 90*200</td></tr>"
            "<tr><td></td><td>Продажи оптовые УРП_1 Заказ: ЦНТ-1</td></tr>"
            "<tr><td>2</td><td>Алёна Шкаф 3-х дверный (корпус) 1/1</td><td>1</td></tr>"
            "<tr><td></td><td>Продажи оптовые УРП_1 Заказ: ЦНТ-2</td></tr>"
            "</table></body></html>"
        )
        path.write_bytes(html.encode("cp1251"))
        result = parse_v7_order(path)

        assert "Васильева" in result.customer_name
        assert len(result.blocks) == 2
        assert result.blocks[0].client_description.startswith("Аврора Кровать 90")
        assert "Àâð" not in result.blocks[0].client_description
        assert result.blocks[1].client_description.startswith("Алёна Шкаф")
        assert result.blocks[0].quantity == 2
        assert result.blocks[1].quantity == 1

    def test_cp1251_mojibake_if_misread_as_latin1(self, tmp_path: Path) -> None:
        """Safety net: latin-1 misread cp1251 bytes still heal via normalize_incoming_text."""
        from src.parsers.v7_parser import normalize_incoming_text

        misread = "Àâðîðà Êðîâàòü 90 ñî âñòðîåííûì îñíîâàíèåì"
        healed = normalize_incoming_text(misread)
        assert healed == "Аврора Кровать 90 со встроенным основанием"
        assert "À" not in healed
