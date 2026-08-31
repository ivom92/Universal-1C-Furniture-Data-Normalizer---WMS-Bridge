"""Soft-furniture transfer detector and hierarchical package parser."""

from __future__ import annotations

from pathlib import Path

from src.adapters.wms_excel_adapter import WMSExcelAdapter
from src.parsers.document_detector import DocumentType, DocumentTypeDetector
from src.parsers.soft_furniture_parser import parse_soft_furniture_order
from src.pipeline import parse_incoming_order, resolve_order_decisions
from tests.excel_fixtures import (
    write_sample_v7_xlsx,
    write_soft_furniture_multi_xlsx,
    write_soft_furniture_xlsx,
)
from tests.test_order_processing import ORDER_BED_CANDIDATES


class TestDocumentTypeDetector:
    def test_standard_picking_list(self, tmp_path: Path) -> None:
        path = write_sample_v7_xlsx(tmp_path / "standard.xlsx")
        assert DocumentTypeDetector().detect(path) == DocumentType.STANDARD_PICKING_LIST

    def test_soft_furniture_markers(self, tmp_path: Path) -> None:
        path = write_soft_furniture_xlsx(tmp_path / "bed.xlsx")
        assert DocumentTypeDetector().detect(path) == DocumentType.SOFT_FURNITURE_TRANSFER


class TestSoftFurnitureParser:
    def test_expands_packages_and_customer(self, tmp_path: Path) -> None:
        path = write_soft_furniture_xlsx(tmp_path / "order_transfering_01_09_bed.xlsx")
        parsed = parse_soft_furniture_order(path)
        assert parsed.customer_name == "ЦНТ-001513"
        assert len(parsed.blocks) == 3
        assert sum(block.quantity for block in parsed.blocks) == 3
        names = [block.client_description for block in parsed.blocks]
        assert all("SF сп Кровать Вена с под мех" in name for name in names)
        assert all("shadow" in name for name in names)
        assert "Изголовье/Дно бельевого короба 1/3" in names[0]
        assert "2/3" in names[1]
        assert "3/3" in names[2]
        assert all(block.quantity == 1 for block in parsed.blocks)

    def test_pipeline_passthrough_wms_columns(self, tmp_path: Path, feature_extractor) -> None:
        from src.matcher.hybrid_matcher import HybridMatcher
        from tests.test_matcher import FakeVectorStore

        path = write_soft_furniture_xlsx(tmp_path / "bed.xlsx")
        doc_type, parsed = parse_incoming_order(path)
        matcher = HybridMatcher(FakeVectorStore([]), feature_extractor)
        decisions = resolve_order_decisions(doc_type, parsed, matcher)
        assert doc_type == DocumentType.SOFT_FURNITURE_TRANSFER
        assert len(decisions) == 3
        assert all(d.status == "MATCHED_AUTO" for d in decisions)
        assert all(d.match_method == "AUTO_NO_BARCODE" for d in decisions)
        assert all(d.matched_entity is None for d in decisions)

        output = tmp_path / "wms.xlsx"
        WMSExcelAdapter().export(decisions, parsed.customer_name, output)
        from openpyxl import load_workbook

        workbook = load_workbook(output)
        assert "Импорт_WMS" in workbook.sheetnames
        assert "Сводка_Отбора" in workbook.sheetnames
        sheet = workbook["Импорт_WMS"]
        assert sheet.cell(2, 3).value in (None, "")
        assert "shadow" in str(sheet.cell(2, 2).value)
        assert sheet.cell(2, 5).value == "ЦНТ-001513"

    def test_multi_parent_packages_and_monoblock(self, tmp_path: Path) -> None:
        path = write_soft_furniture_multi_xlsx(tmp_path / "multi.xlsx")
        parsed = parse_soft_furniture_order(path)
        assert len(parsed.blocks) == 6
        assert sum(block.quantity for block in parsed.blocks) == 6
        names = [block.client_description for block in parsed.blocks]
        orders = [block.order_service_line for block in parsed.blocks]
        assert names[0] == "Диван Милан velvet Сиденье 1/2"
        assert names[1] == "Диван Милан velvet Спинка 2/2"
        assert names[2] == "Кровать Вена loft Изголовье 1/3"
        assert names[3] == "Кровать Вена loft Боковины 2/3"
        assert names[4] == "Кровать Вена loft Основание 3/3"
        assert names[5] == "Пуф Куб SF"
        assert orders[:2] == ["ЦНТ-100", "ЦНТ-100"]
        assert orders[2:5] == ["ЦНТ-200", "ЦНТ-200", "ЦНТ-200"]
        assert orders[5] == "ЦНТ-300"

    def test_real_bed_file_three_places(self) -> None:
        path = next((item for item in ORDER_BED_CANDIDATES if item.exists()), None)
        if path is None:
            import pytest

            pytest.skip("Soft-furniture transfer sample is missing")
        parsed = parse_soft_furniture_order(path)
        assert len(parsed.blocks) == 3
        assert sum(block.quantity for block in parsed.blocks) == 3
