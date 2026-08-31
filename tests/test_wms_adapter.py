"""Production-grade checks for the WMS Excel adapter."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.adapters.wms_excel_adapter import (
    NO_BARCODE_INSTRUCTION,
    STATUS_NO_FACTORY_BARCODE,
    STATUS_OPERATOR_SCAN,
    STATUS_QUARANTINE_PICK,
    SUMMARY_SHEET_NAME,
    WMS_COLUMNS,
    WMS_SHEET_NAME,
    WMSExcelAdapter,
    auto_filter_data_last_row,
)
from src.models import CatalogEntity, ExtractedFeatures, MatchDecision, RawOrderBlock


def _block(*, description: str, quantity: int = 2, line_number: int = 1) -> RawOrderBlock:
    return RawOrderBlock(
        line_number=line_number,
        client_description=description,
        item_type="Пачка",
        quantity=quantity,
        factory_alias=description,
        order_service_line="Продажи оптовые УРП_ test",
        excel_row_start=line_number,
    )


def _entity(*, nomenclature: str, barcode: str | None) -> CatalogEntity:
    return CatalogEntity.model_validate(
        {
            "Номенклатура": nomenclature,
            "НоменклатураКод": "00000010001",
            "Штрихкод": barcode,
            "Упаковка": "1/1",
        }
    )


def _decision(
    *,
    description: str,
    quantity: int,
    status: str,
    entity: CatalogEntity | None,
    line_number: int = 1,
    match_method: str | None = None,
    status_detail: str | None = None,
) -> MatchDecision:
    if match_method is None:
        match_method = "vector_auto" if entity is not None else "AUTO_NO_BARCODE"
    return MatchDecision(
        raw_block=_block(description=description, quantity=quantity, line_number=line_number),
        extracted_features=ExtractedFeatures(),
        status=status,
        matched_entity=entity,
        confidence_score=1.0,
        match_method=match_method,
        status_detail=status_detail,
    )


def _sample_decisions() -> list[MatchDecision]:
    return [
        _decision(
            description="Клиентское имя 1",
            quantity=2,
            status="MATCHED_AUTO",
            entity=_entity(nomenclature="Фабрика Кухня Равенна Н20 1/1", barcode="2006000045445"),
            line_number=1,
        ),
        _decision(
            description="Фурнитура Полка стеклянная 30/40/50/60/80 (полкодержатели 8 шт)",
            quantity=3,
            status="MATCHED_AUTO",
            entity=None,
            line_number=2,
        ),
        _decision(
            description="Стекло 565х255 заказное",
            quantity=4,
            status="QUARANTINE",
            entity=None,
            line_number=3,
            match_method="QUARANTINE",
            status_detail="Нестандартный заказной размер",
        ),
    ]


class TestWMSExcelAdapter:
    def test_wms_export_column_formats(self, tmp_path: Path) -> None:
        path = tmp_path / "wms.xlsx"
        WMSExcelAdapter().export(_sample_decisions(), "Тестовый заказчик", path)
        worksheet = load_workbook(path).active
        number_cell = worksheet.cell(row=2, column=1)
        barcode_cell = worksheet.cell(row=2, column=3)
        qty_cell = worksheet.cell(row=2, column=4)
        assert number_cell.value == 1
        assert isinstance(number_cell.value, int)
        assert number_cell.number_format == "0"
        assert barcode_cell.number_format == "@"
        assert barcode_cell.value == "2006000045445"
        assert isinstance(qty_cell.value, int)
        assert qty_cell.value == 2
        assert qty_cell.number_format == "0"
        empty_barcode = worksheet.cell(row=3, column=3)
        assert empty_barcode.number_format == "@"

    def test_wms_export_no_string_none(self, tmp_path: Path) -> None:
        path = tmp_path / "wms.xlsx"
        decisions = _sample_decisions()
        WMSExcelAdapter().export(decisions, "Тестовый заказчик", path)
        worksheet = load_workbook(path).active
        data_last = WMSExcelAdapter.wms_data_last_row(len(decisions))
        for row_index in range(2, data_last + 1):
            value = worksheet.cell(row=row_index, column=3).value
            assert value not in {"None", "nan", "null", "none"}
            if row_index > 2:
                assert value in (None, "")

    def test_wms_export_autofit_and_freeze(self, tmp_path: Path) -> None:
        path = tmp_path / "wms.xlsx"
        WMSExcelAdapter().export(_sample_decisions(), "Тестовый заказчик ИП", path)
        worksheet = load_workbook(path).active
        assert worksheet.freeze_panes == "A2"
        assert 8 <= worksheet.column_dimensions["A"].width <= 10
        assert worksheet.column_dimensions["B"].width >= 50
        assert worksheet.column_dimensions["C"].width >= 18
        assert worksheet.column_dimensions["D"].width >= 12
        assert worksheet.column_dimensions["E"].width >= 25
        header = worksheet.cell(row=1, column=1)
        assert header.value == "№"
        assert header.font.bold
        assert header.font.color.rgb.endswith("FFFFFF")
        assert header.fill.fgColor.rgb.endswith("2F3542")
        assert worksheet.row_dimensions[1].height == 26
        assert [worksheet.cell(row=1, column=col).value for col in range(1, 6)] == WMS_COLUMNS

    def test_wms_export_zero_loss_row_count(self, tmp_path: Path) -> None:
        decisions = _sample_decisions()
        path = tmp_path / "wms.xlsx"
        WMSExcelAdapter().export(decisions, "Тестовый заказчик", path)
        worksheet = load_workbook(path).active
        data_last = WMSExcelAdapter.wms_data_last_row(len(decisions))
        data_rows = data_last - 1
        assert data_rows == len(decisions)
        qty_sum = sum(int(worksheet.cell(row=row, column=4).value) for row in range(2, data_last + 1))
        assert qty_sum == sum(item.raw_block.quantity for item in decisions)
        assert worksheet.cell(row=2, column=2).value == "Фабрика Кухня Равенна Н20 1/1"
        assert worksheet.cell(row=3, column=2).value.startswith("Фурнитура Полка стеклянная")
        assert worksheet.cell(row=4, column=2).value == "Стекло 565х255 заказное"

    def test_wms_export_strict_line_number_order(self, tmp_path: Path) -> None:
        shuffled = [
            _decision(
                description="Третья",
                quantity=4,
                status="QUARANTINE",
                entity=None,
                line_number=3,
            ),
            _decision(
                description="Первая",
                quantity=1,
                status="MATCHED_AUTO",
                entity=_entity(nomenclature="Имя 1", barcode="2006000000001"),
                line_number=1,
            ),
            _decision(
                description="Вторая",
                quantity=2,
                status="MATCHED_AUTO",
                entity=None,
                line_number=2,
            ),
        ]
        path = tmp_path / "wms_shuffled.xlsx"
        WMSExcelAdapter().export(shuffled, "Заказчик", path)
        worksheet = load_workbook(path).active
        numbers = [worksheet.cell(row=row, column=1).value for row in range(2, 5)]
        names = [worksheet.cell(row=row, column=2).value for row in range(2, 5)]
        assert numbers == [1, 2, 3]
        assert all(isinstance(value, int) and not isinstance(value, bool) for value in numbers)
        assert names == ["Имя 1", "Вторая", "Третья"]

    def test_wms_export_totals_row_formula_and_autofilter(self, tmp_path: Path) -> None:
        decisions = _sample_decisions()
        path = tmp_path / "wms.xlsx"
        WMSExcelAdapter().export(decisions, "Тестовый заказчик", path, source_name="order.xls")
        workbook = load_workbook(path)
        worksheet = workbook[WMS_SHEET_NAME]
        n = len(decisions)
        data_last = WMSExcelAdapter.wms_data_last_row(n)
        totals_row = WMSExcelAdapter.wms_totals_row(n)
        assert worksheet.max_row == totals_row
        assert worksheet.auto_filter.ref == f"A1:E{data_last}"
        assert auto_filter_data_last_row(worksheet) == data_last
        name_cell = worksheet.cell(row=totals_row, column=2)
        qty_cell = worksheet.cell(row=totals_row, column=4)
        assert name_cell.value == f"ИТОГО (Позиций: {n})"
        assert name_cell.font.bold
        assert str(qty_cell.value) == f"=SUM(D2:D{data_last})"
        assert qty_cell.font.bold
        assert worksheet.cell(row=totals_row, column=1).value == "Σ"
        assert worksheet.cell(row=totals_row, column=3).value in (None, "")
        assert worksheet.cell(row=totals_row, column=5).value in (None, "")
        assert name_cell.fill.fgColor.rgb.endswith("F1F2F6")
        assert qty_cell.border.bottom.style == "double"
        assert qty_cell.border.top.style == "thin"

    def test_wms_export_has_summary_sheet(self, tmp_path: Path) -> None:
        decisions = _sample_decisions()
        path = tmp_path / "wms.xlsx"
        WMSExcelAdapter().export(
            decisions,
            "Тестовый заказчик",
            path,
            source_name="order_transfering_01_09.xls",
        )
        workbook = load_workbook(path)
        assert workbook.sheetnames == [WMS_SHEET_NAME, SUMMARY_SHEET_NAME]
        summary = workbook[SUMMARY_SHEET_NAME]
        assert "Сводка отбора" in str(summary.cell(row=1, column=1).value)
        assert summary.cell(row=3, column=2).value == "Тестовый заказчик"
        assert summary.cell(row=4, column=2).value == "order_transfering_01_09.xls"
        blob = " ".join(
            str(summary.cell(row=row, column=col).value or "")
            for row in range(1, summary.max_row + 1)
            for col in range(1, 5)
        )
        assert "Нестандартный заказной размер" in blob
        assert "Карантин" in blob
        assert "Фурнитура Полка стеклянная" in blob
        assert "AUTO_NO_BARCODE" not in blob
        assert STATUS_NO_FACTORY_BARCODE in blob
        assert STATUS_QUARANTINE_PICK in blob
        assert "заводского штрихкода" in blob
        assert NO_BARCODE_INSTRUCTION[:40] in blob

    def test_wms_export_with_operator_overrides(self, tmp_path: Path) -> None:
        decisions = _sample_decisions()
        scanned = "2006000099999"
        path = tmp_path / "wms_overrides.xlsx"
        WMSExcelAdapter().export(
            decisions,
            "Тестовый заказчик",
            path,
            source_name="order.xls",
            overrides={2: scanned},
        )
        workbook = load_workbook(path)
        wms = workbook[WMS_SHEET_NAME]
        assert wms.cell(row=3, column=3).value == scanned
        assert wms.cell(row=3, column=3).number_format == "@"
        assert isinstance(wms.cell(row=3, column=3).value, str)
        summary = workbook[SUMMARY_SHEET_NAME]
        blob = " ".join(
            str(summary.cell(row=row, column=col).value or "")
            for row in range(1, summary.max_row + 1)
            for col in range(1, 5)
        )
        assert "AUTO_NO_BARCODE" not in blob
        assert STATUS_OPERATOR_SCAN in blob
        assert scanned not in {None, ""}
        assert WMSExcelAdapter.count_matched_with_barcode(decisions, {2: scanned}) == 2
